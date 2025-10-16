#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Скрипт для генерации Excel файла с матрицей доступа к страницам Check_Site
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Создаем новую рабочую книгу
wb = Workbook()
ws = wb.active
ws.title = "Матрица доступа"

# Роли пользователей
roles = [
    'SUPERADMIN',
    'DIRECTOR',
    'CHIEF_ENGINEER',
    'PROJECT_MANAGER',
    'ENGINEER',
    'SITE_MANAGER',
    'FOREMAN',
    'MASTER',
    'SUPERVISOR',
    'CONTRACTOR',
    'OBSERVER'
]

# Русские названия ролей
roles_ru = {
    'SUPERADMIN': 'Суперадмин',
    'DIRECTOR': 'Директор',
    'CHIEF_ENGINEER': 'Главный инженер',
    'PROJECT_MANAGER': 'Руководитель проекта',
    'ENGINEER': 'Инженер ПТО',
    'SITE_MANAGER': 'Начальник участка',
    'FOREMAN': 'Прораб',
    'MASTER': 'Мастер',
    'SUPERVISOR': 'Технадзор',
    'CONTRACTOR': 'Подрядчик',
    'OBSERVER': 'Наблюдатель'
}

# Страницы и их доступность
pages_access = [
    ('Дашборд', [True] * 11),
    ('Проекты', [True] * 11),
    ('Замечания', [True] * 11),
    ('Пользователи', [True, True, True, True, True, True, False, False, False, False, False]),
    ('Подрядчики', [True, True, True, True, True, True, False, False, False, False, False]),
    ('Надзоры', [True, True, True, True, True, True, True, False, False, False, False]),
    ('Техусловия', [True] * 11),
    ('Отчеты', [True, True, True, True, True, True, True, True, True, False, True]),
]

# Цвета
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True, size=11)
yes_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
no_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
page_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

# Границы
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Заголовок документа
ws.merge_cells('A1:L1')
title_cell = ws['A1']
title_cell.value = 'МАТРИЦА ДОСТУПА К СТРАНИЦАМ - Check_Site'
title_cell.font = Font(bold=True, size=14, color='1F4E78')
title_cell.alignment = Alignment(horizontal='center', vertical='center')

# Заголовки столбцов (строка 3)
ws['A3'] = 'Страница'
ws['A3'].font = header_font
ws['A3'].fill = header_fill
ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
ws['A3'].border = thin_border

for idx, role in enumerate(roles, start=2):
    col_letter = get_column_letter(idx)
    cell = ws[f'{col_letter}3']
    cell.value = roles_ru[role]
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Заполняем данные
row = 4
for page_name, access_list in pages_access:
    # Название страницы
    ws[f'A{row}'] = page_name
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'A{row}'].fill = page_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws[f'A{row}'].border = thin_border

    # Доступ для каждой роли
    for idx, has_access in enumerate(access_list, start=2):
        col_letter = get_column_letter(idx)
        cell = ws[f'{col_letter}{row}']
        cell.value = '✅' if has_access else '❌'
        cell.fill = yes_fill if has_access else no_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        cell.font = Font(size=14)

    row += 1

# Добавляем примечание для страницы "Пользователи"
note_row = row + 1
ws.merge_cells(f'A{note_row}:L{note_row}')
note_cell = ws[f'A{note_row}']
note_cell.value = '* Для страницы "Пользователи" требуется дополнительно approved=true'
note_cell.font = Font(italic=True, size=10, color='666666')
note_cell.alignment = Alignment(horizontal='left', vertical='center')

# Легенда
legend_row = note_row + 2
ws.merge_cells(f'A{legend_row}:L{legend_row}')
legend_cell = ws[f'A{legend_row}']
legend_cell.value = 'Легенда:'
legend_cell.font = Font(bold=True, size=11)

legend_row += 1
ws[f'A{legend_row}'] = '✅ - Доступ разрешен'
ws[f'A{legend_row}'].fill = yes_fill
ws[f'A{legend_row}'].border = thin_border

ws[f'D{legend_row}'] = '❌ - Доступ запрещен'
ws[f'D{legend_row}'].fill = no_fill
ws[f'D{legend_row}'].border = thin_border

# Настройка ширины столбцов
ws.column_dimensions['A'].width = 20
for idx in range(2, 13):
    col_letter = get_column_letter(idx)
    ws.column_dimensions[col_letter].width = 15

# Настройка высоты строк
ws.row_dimensions[1].height = 25
ws.row_dimensions[3].height = 35

# Сохраняем файл
output_file = '/Users/kairatkhidirboev/Projects/checksite/Матрица_доступа_страниц.xlsx'
wb.save(output_file)

print(f'✅ Excel файл успешно создан: {output_file}')
