"""
Утилиты для работы с Excel файлами для импорта/экспорта проектов.
"""
from io import BytesIO
from typing import List, Dict, Any
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


def generate_excel_template() -> HttpResponse:
    """
    Генерирует шаблон Excel для импорта проектов.

    Структура шаблона:
    - Наименование объекта* (обязательное)
    - Страна (необязательное)
    - Адрес* (обязательное)

    Returns:
        HttpResponse с Excel файлом
    """
    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Проекты"

    # Заголовки столбцов
    headers = [
        'Наименование объекта*',
        'Страна',
        'Адрес*'
    ]

    # Записываем заголовки
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        # Стиль заголовка: жирный шрифт, синий фон
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Примеры данных
    examples = [
        ['Жилой комплекс "Алатау"', 'Казахстан', 'ул. Абая 150'],
        ['Торговый центр "Mega"', 'Казахстан', 'ул. Розыбакиева 247'],
    ]

    for row_num, example in enumerate(examples, 2):
        for col_num, value in enumerate(example, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Автоматическая ширина столбцов
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 35

    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Формируем HTTP ответ
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="projects_template.xlsx"'

    return response


def parse_excel_file(file) -> Dict[str, Any]:
    """
    Парсит загруженный Excel файл и валидирует данные.

    Args:
        file: Загруженный файл (UploadedFile объект)

    Returns:
        Dict с ключами:
            - 'success': bool - успешность парсинга
            - 'data': List[Dict] - список проектов для создания
            - 'errors': List[str] - список ошибок валидации

    Raises:
        Exception: Если файл не может быть прочитан
    """
    result = {
        'success': True,
        'data': [],
        'errors': []
    }

    try:
        # Загружаем Excel файл
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active

        # Получаем заголовки из первой строки
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())

        # Проверяем наличие обязательных колонок
        required_headers = ['Наименование объекта*', 'Адрес*']
        missing_headers = []

        for req_header in required_headers:
            # Ищем заголовок (с учетом возможных вариаций)
            base_header = req_header.replace('*', '').strip()
            found = False
            for header in headers:
                if base_header.lower() in header.lower():
                    found = True
                    break
            if not found:
                missing_headers.append(req_header)

        if missing_headers:
            result['success'] = False
            result['errors'].append(
                f'Отсутствуют обязательные столбцы: {", ".join(missing_headers)}'
            )
            return result

        # Определяем индексы столбцов
        name_col_idx = None
        country_col_idx = None
        address_col_idx = None

        for idx, header in enumerate(headers):
            header_lower = header.lower()
            if 'наименование' in header_lower or 'название' in header_lower:
                name_col_idx = idx
            elif 'страна' in header_lower:
                country_col_idx = idx
            elif 'адрес' in header_lower:
                address_col_idx = idx

        # Парсим данные (начиная со второй строки)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Пропускаем пустые строки
            if not any(row):
                continue

            # Извлекаем значения
            name = row[name_col_idx].strip() if name_col_idx is not None and row[name_col_idx] else None
            country = row[country_col_idx].strip() if country_col_idx is not None and len(row) > country_col_idx and row[country_col_idx] else ''
            address = row[address_col_idx].strip() if address_col_idx is not None and row[address_col_idx] else None

            # Валидация обязательных полей
            row_errors = []
            if not name:
                row_errors.append('Наименование объекта')
            if not address:
                row_errors.append('Адрес')

            if row_errors:
                result['errors'].append(
                    f'Строка {row_num}: отсутствуют обязательные поля: {", ".join(row_errors)}'
                )
                result['success'] = False
                continue

            # Добавляем валидированные данные
            result['data'].append({
                'name': name,
                'address': address,
                'customer': country if country else ''  # Используем поле customer для страны
            })

        # Проверяем, есть ли хоть один проект
        if not result['data'] and result['success']:
            result['success'] = False
            result['errors'].append('Файл не содержит данных для импорта')

        wb.close()

    except Exception as e:
        result['success'] = False
        result['errors'].append(f'Ошибка при чтении файла: {str(e)}')

    return result


def generate_excel_export(projects: List) -> HttpResponse:
    """
    Генерирует Excel файл с экспортом всех проектов компании.

    Структура экспорта:
    - Наименование объекта
    - Страна
    - Адрес
    - Закреплённые сотрудники
    - Надзоры
    - Подрядчики

    Args:
        projects: QuerySet или список объектов Project

    Returns:
        HttpResponse с Excel файлом
    """
    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Проекты"

    # Заголовки столбцов
    headers = [
        'Наименование объекта',
        'Страна',
        'Адрес',
        'Закреплённые сотрудники',
        'Надзоры',
        'Подрядчики'
    ]

    # Записываем заголовки
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        # Стиль заголовка
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Записываем данные проектов
    for row_num, project in enumerate(projects, start=2):
        # Наименование объекта
        ws.cell(row=row_num, column=1, value=project.name)

        # Страна (используем поле customer)
        ws.cell(row=row_num, column=2, value=project.customer if project.customer else '')

        # Адрес
        ws.cell(row=row_num, column=3, value=project.address)

        # Получаем всех участников проекта из ManyToMany поля team_members
        all_team_members = project.team_members.all()

        # Закреплённые сотрудники (все роли КРОМЕ CONTRACTOR и SUPERVISOR)
        # Фильтруем сотрудников: исключаем подрядчиков и надзоры
        staff_members = [m for m in all_team_members if m.role not in ['CONTRACTOR', 'SUPERVISOR']]
        staff_str = ', '.join([member.get_full_name() for member in staff_members])
        ws.cell(row=row_num, column=4, value=staff_str)

        # Надзоры (только роль SUPERVISOR - Технадзор)
        # Фильтруем участников с ролью 'SUPERVISOR'
        supervisors = [m for m in all_team_members if m.role == 'SUPERVISOR']
        supervisors_str = ', '.join([sup.get_full_name() for sup in supervisors])
        ws.cell(row=row_num, column=5, value=supervisors_str)

        # Подрядчики (только роль CONTRACTOR - Подрядчик)
        # Фильтруем участников с ролью 'CONTRACTOR'
        contractors = [m for m in all_team_members if m.role == 'CONTRACTOR']
        contractors_str = ', '.join([contr.get_full_name() for contr in contractors])
        ws.cell(row=row_num, column=6, value=contractors_str)

        # Выравнивание текста для всех ячеек строки
        # Применяем стиль: выравнивание влево, по верхнему краю, с переносом текста
        for col_num in range(1, 7):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Автоматическая ширина столбцов
    column_widths = [40, 20, 40, 30, 30, 30]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Формируем HTTP ответ
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="projects_export.xlsx"'

    return response
