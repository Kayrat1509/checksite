from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import ValidationError
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from django.db.models import Q
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

from .models import (
    MaterialRequest,
    MaterialRequestItem,
    MaterialRequestDocument,
    MaterialRequestComment
)
from .serializers import (
    MaterialRequestListSerializer,
    MaterialRequestDetailSerializer,
    MaterialRequestCreateSerializer,
    MaterialRequestUpdateSerializer,
    MaterialRequestStatusChangeSerializer,
    MaterialRequestItemSerializer,
    MaterialRequestDocumentSerializer,
    MaterialRequestCommentSerializer
)
from .permissions import (
    MaterialRequestPermission,
    MaterialRequestStatusChangePermission,
    MaterialRequestDocumentPermission,
    MaterialRequestCommentPermission
)
from apps.core.viewsets import SoftDeleteViewSetMixin


class MaterialRequestViewSet(SoftDeleteViewSetMixin, viewsets.ModelViewSet):
    """
    ViewSet для работы с заявками на материалы.

    Endpoints согласно ТЗ (раздел 5):
    - GET /api/material-requests/ - список заявок
    - GET /api/material-requests/{id}/ - детальная карточка
    - POST /api/material-requests/ - создать заявку
    - PUT/PATCH /api/material-requests/{id}/ - редактировать заявку
    - DELETE /api/material-requests/{id}/ - удалить заявку
    - PATCH /api/material-requests/{id}/status/ - смена статуса
    - POST /api/material-requests/{id}/upload/ - загрузка файлов
    """

    permission_classes = [IsAuthenticated, MaterialRequestPermission]
    pagination_class = None  # Отключаем пагинацию, возвращаем все заявки
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['project', 'author', 'status', 'responsible']
    search_fields = ['request_number', 'notes', 'items__material_name']
    ordering_fields = ['created_at', 'updated_at', 'request_number']
    ordering = ['-created_at']

    def get_queryset(self):
        """
        Получение списка заявок с учетом прав доступа.
        Согласно ТЗ: каждый сотрудник видит заявки своих объектов.
        """
        user = self.request.user

        # Суперадмин и руководители видят все заявки
        if user.is_superuser or user.role in ['SUPERADMIN', 'DIRECTOR', 'CHIEF_ENGINEER']:
            return MaterialRequest.objects.all().select_related(
                'project',
                'author',
                'responsible'
            ).prefetch_related('items', 'documents', 'history', 'comments')

        # Остальные видят заявки своих проектов
        user_projects = user.projects.all() if hasattr(user, 'projects') else []

        return MaterialRequest.objects.filter(
            Q(project__in=user_projects) |
            Q(author=user) |
            Q(responsible=user)
        ).select_related(
            'project',
            'author',
            'responsible'
        ).prefetch_related('items', 'documents', 'history', 'comments').distinct()

    def get_serializer_class(self):
        """Выбор сериализатора в зависимости от действия."""
        if self.action == 'list':
            return MaterialRequestListSerializer
        elif self.action == 'create':
            return MaterialRequestCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return MaterialRequestUpdateSerializer
        elif self.action == 'change_status':
            return MaterialRequestStatusChangeSerializer
        return MaterialRequestDetailSerializer

    def perform_create(self, serializer):
        """Создание заявки с установкой автора."""
        serializer.save()

    @action(detail=True, methods=['patch'], permission_classes=[IsAuthenticated, MaterialRequestStatusChangePermission])
    def change_status(self, request, pk=None):
        """
        Изменение статуса заявки.
        Endpoint: PATCH /api/material-requests/{id}/status/

        Body: {
            "new_status": "UNDER_REVIEW",
            "comment": "Отправлено на проверку"
        }
        """
        material_request = self.get_object()
        serializer = self.get_serializer(
            data=request.data,
            context={
                'material_request': material_request,
                'request': request
            }
        )
        serializer.is_valid(raise_exception=True)
        updated_request = serializer.save()

        # Возвращаем обновленную заявку
        detail_serializer = MaterialRequestDetailSerializer(
            updated_request,
            context={'request': request}
        )
        return Response(detail_serializer.data)

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated, MaterialRequestDocumentPermission])
    def upload(self, request, pk=None):
        """
        Загрузка документов к заявке.
        Endpoint: POST /api/material-requests/{id}/upload/

        Multipart form data:
        - document_type: тип документа (REGISTRY, INVOICE, WAYBILL, PHOTO, OTHER)
        - file: файл
        - file_name: название файла
        - description: описание (опционально)
        """
        material_request = self.get_object()

        serializer = MaterialRequestDocumentSerializer(
            data=request.data,
            context={'request': request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save(
            request=material_request,
            uploaded_by=request.user
        )

        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated, MaterialRequestCommentPermission])
    def add_comment(self, request, pk=None):
        """
        Добавление комментария к заявке.
        Endpoint: POST /api/material-requests/{id}/add_comment/

        Body: {
            "text": "Текст комментария"
        }
        """
        material_request = self.get_object()

        serializer = MaterialRequestCommentSerializer(
            data=request.data,
            context={'request': request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save(
            request=material_request,
            author=request.user
        )

        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def add_item(self, request, pk=None):
        """
        Добавление позиции материала к заявке.
        Endpoint: POST /api/material-requests/{id}/add_item/

        Body: {
            "material_name": "Арматура Ø12",
            "quantity": 1500,
            "unit": "кг",
            "specifications": "Описание",
            "order": 1
        }
        """
        material_request = self.get_object()

        # Проверка прав - только автор или ответственный
        if (request.user != material_request.author and
            request.user != material_request.responsible and
            request.user.role not in ['SUPERADMIN', 'DIRECTOR', 'CHIEF_ENGINEER']):
            return Response(
                {'detail': 'У вас нет прав для добавления материалов к этой заявке'},
                status=status.HTTP_403_FORBIDDEN
            )

        serializer = MaterialRequestItemSerializer(
            data=request.data,
            context={'request': request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save(request=material_request)

        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['get'])
    def export_excel(self, request, pk=None):
        """
        Экспорт заявки в Excel.
        Endpoint: GET /api/material-requests/{id}/export_excel/

        Возвращает файл Excel с полной информацией о заявке.
        """
        material_request = self.get_object()

        # Создаем новую книгу Excel
        wb = Workbook()
        ws = wb.active
        # Убираем недопустимые символы из названия листа (Excel не поддерживает: / \ ? * [ ])
        safe_number = material_request.request_number.replace('/', '-')
        ws.title = f"Заявка {safe_number}"

        # Определяем стили
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=11)
        regular_font = Font(size=11)

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Заголовок документа
        ws.merge_cells('A1:F1')
        ws['A1'] = f'ЗАЯВКА НА МАТЕРИАЛЫ № {material_request.request_number}'
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = center_alignment
        ws['A1'].fill = header_fill
        ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")

        # Информация о заявке
        row = 3
        ws[f'A{row}'] = 'Дата создания:'
        ws[f'B{row}'] = material_request.created_at.strftime('%d.%m.%Y %H:%M')
        ws[f'A{row}'].font = subheader_font

        row += 1
        ws[f'A{row}'] = 'Объект:'
        ws[f'B{row}'] = material_request.project.name if material_request.project else '-'
        ws[f'A{row}'].font = subheader_font

        row += 1
        ws[f'A{row}'] = 'Автор (Прораб):'
        ws[f'B{row}'] = material_request.author.get_full_name() if material_request.author else '-'
        ws[f'A{row}'].font = subheader_font

        row += 1
        ws[f'A{row}'] = 'Статус:'
        ws[f'B{row}'] = material_request.get_status_display()
        ws[f'A{row}'].font = subheader_font

        if material_request.responsible:
            row += 1
            ws[f'A{row}'] = 'Ответственный:'
            ws[f'B{row}'] = material_request.responsible.get_full_name()
            ws[f'A{row}'].font = subheader_font

        if material_request.drawing_reference:
            row += 1
            ws[f'A{row}'] = 'Чертёж / Лист:'
            ws[f'B{row}'] = material_request.drawing_reference
            ws[f'A{row}'].font = subheader_font

        if material_request.work_type:
            row += 1
            ws[f'A{row}'] = 'Вид работ:'
            ws[f'B{row}'] = material_request.work_type
            ws[f'A{row}'].font = subheader_font

        if material_request.notes:
            row += 1
            ws[f'A{row}'] = 'Примечание:'
            ws[f'B{row}'] = material_request.notes
            ws[f'A{row}'].font = subheader_font

        # Таблица материалов
        row += 2
        ws[f'A{row}'] = '№'
        ws[f'B{row}'] = 'Наименование материала'
        ws[f'C{row}'] = 'Количество'
        ws[f'D{row}'] = 'Ед. изм.'
        ws[f'E{row}'] = 'Примечания'
        ws[f'F{row}'] = 'Дата добавления'

        # Применяем стиль к заголовкам таблицы
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = ws[f'{col}{row}']
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border

        # Заполняем данные материалов
        items = material_request.items.order_by('order')
        for idx, item in enumerate(items, 1):
            row += 1
            ws[f'A{row}'] = idx
            ws[f'B{row}'] = item.material_name
            ws[f'C{row}'] = float(item.quantity)
            ws[f'D{row}'] = item.unit
            ws[f'E{row}'] = item.specifications or '-'
            ws[f'F{row}'] = item.created_at.strftime('%d.%m.%Y')

            # Применяем стиль к ячейкам
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                cell = ws[f'{col}{row}']
                cell.border = border
                cell.font = regular_font
                if col == 'A':
                    cell.alignment = center_alignment
                elif col == 'C':
                    cell.alignment = center_alignment
                else:
                    cell.alignment = left_alignment

        # Настраиваем ширину колонок
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 15

        # Подпись в конце документа
        row += 3
        ws[f'A{row}'] = f'Всего позиций: {items.count()}'
        ws[f'A{row}'].font = subheader_font

        # Создаем HTTP-ответ с Excel файлом
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'Заявка_{material_request.request_number}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response


class MaterialRequestItemViewSet(viewsets.ModelViewSet):
    """ViewSet для работы с позициями материалов."""

    queryset = MaterialRequestItem.objects.all().select_related('request', 'cancelled_by', 'issued_by', 'request__project')
    serializer_class = MaterialRequestItemSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['request', 'request__project', 'unit', 'status']
    search_fields = ['material_name', 'specifications']

    def get_queryset(self):
        """Кастомный queryset с возможностью фильтрации по actual_quantity."""
        queryset = super().get_queryset()

        # Фильтр: только материалы с заполненным actual_quantity
        has_actual_quantity = self.request.query_params.get('has_actual_quantity')
        if has_actual_quantity and has_actual_quantity.lower() == 'true':
            queryset = queryset.filter(actual_quantity__isnull=False)
            # Исключаем отмененные позиции для страницы Склад
            queryset = queryset.exclude(status='CANCELLED')

        return queryset

    def update(self, request, *args, **kwargs):
        """
        Обновление позиции материала.
        Разрешено только автору при статусе RETURNED_FOR_REVISION или DRAFT.
        Исключение: поле actual_quantity может обновлять автор в любом статусе.
        Исключение: поле issued_quantity могут обновлять Завсклад объекта и Начальник участка.
        """
        item = self.get_object()
        material_request = item.request
        user = request.user

        # Логируем данные запроса для отладки
        print(f"DEBUG: Updating item {item.id}, request.data = {request.data}, user.role = {user.role}")

        # Если обновляется только actual_quantity, разрешаем автору в любом статусе
        is_only_actual_quantity = (
            len(request.data) == 1 and 'actual_quantity' in request.data
        )

        # Если обновляется только issued_quantity, разрешаем Завскладу объекта и Начальнику участка
        is_only_issued_quantity = (
            len(request.data) == 1 and 'issued_quantity' in request.data
        )

        # Проверка прав для issued_quantity
        if is_only_issued_quantity:
            # Только Завсклад объекта и Начальник участка могут редактировать issued_quantity
            if user.role not in ['SITE_WAREHOUSE_MANAGER', 'FOREMAN']:
                print(f"DEBUG: User role {user.role} not allowed to edit issued_quantity")
                return Response(
                    {'detail': 'Редактирование количества выдано доступно только для Завсклада объекта и Начальника участка'},
                    status=status.HTTP_403_FORBIDDEN
                )
            # Пропускаем проверку - разрешаем обновление

        # Проверка прав: автор может редактировать при RETURNED_FOR_REVISION или DRAFT
        elif user.role == 'SUPERADMIN':
            pass  # Суперадмин может всё
        elif material_request.author == user:
            # Если обновляется только actual_quantity, разрешаем в любом статусе
            if is_only_actual_quantity:
                pass  # Разрешаем обновление actual_quantity
            # Остальные поля только при определенных статусах
            elif item.item_status not in ['DRAFT', 'RETURNED_FOR_REVISION']:
                return Response(
                    {'detail': 'Редактирование позиции доступно только в статусах "Черновик" или "На доработке"'},
                    status=status.HTTP_403_FORBIDDEN
                )
        else:
            return Response(
                {'detail': 'У вас нет прав для редактирования этой позиции'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Вызываем родительский метод и логируем результат
        try:
            response = super().update(request, *args, **kwargs)
            print(f"DEBUG: Update successful for item {item.id}")
            return response
        except ValidationError as e:
            print(f"DEBUG: Validation error for item {item.id}: {e.detail}")
            raise
        except Exception as e:
            print(f"DEBUG: Update failed for item {item.id}, error: {type(e).__name__}: {str(e)}")
            raise

    def perform_update(self, serializer):
        """
        Переопределяем perform_update для автоматического заполнения issued_by.
        """
        # Если обновляется issued_quantity, автоматически заполняем issued_by
        if 'issued_quantity' in self.request.data:
            serializer.save(issued_by=self.request.user)
        else:
            serializer.save()

    def partial_update(self, request, *args, **kwargs):
        """Частичное обновление позиции материала (PATCH)."""
        kwargs['partial'] = True
        return self.update(request, *args, **kwargs)

    @action(detail=True, methods=['patch'])
    def cancel_item(self, request, pk=None):
        """
        Отмена позиции материала.
        Endpoint: PATCH /api/material-request-items/{id}/cancel_item/

        Body: {
            "cancellation_reason": "Причина отмены"
        }
        """
        from django.utils import timezone

        item = self.get_object()
        material_request = item.request

        # Проверка прав - только автор или ответственный или суперадмин
        if (request.user != material_request.author and
            request.user != material_request.responsible and
            request.user.role not in ['SUPERADMIN', 'DIRECTOR', 'CHIEF_ENGINEER']):
            return Response(
                {'detail': 'У вас нет прав для отмены этой позиции'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Проверка, что позиция не была отменена ранее
        if item.status == 'CANCELLED':
            return Response(
                {'detail': 'Позиция уже отменена'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Получаем причину отмены из запроса
        cancellation_reason = request.data.get('cancellation_reason', 'Отменено без указания причины')

        # Сохраняем текущий статус позиции для возможности восстановления
        item.previous_item_status = item.item_status

        # Обновляем статус позиции
        item.status = 'CANCELLED'
        item.cancellation_reason = cancellation_reason
        item.cancelled_by = request.user
        item.cancelled_at = timezone.now()
        item.save()

        # Возвращаем обновленную позицию
        serializer = self.get_serializer(item)
        return Response(serializer.data)

    @action(detail=True, methods=['patch'])
    def restore_item(self, request, pk=None):
        """
        Восстановление отмененной позиции материала.
        Позиция возвращается к статусу, который был до отмены (продолжает согласование).

        Endpoint: PATCH /api/material-request-items/{id}/restore_item/
        """
        from django.utils import timezone

        item = self.get_object()
        material_request = item.request

        # Проверка прав - только автор или ответственный или суперадмин
        if (request.user != material_request.author and
            request.user != material_request.responsible and
            request.user.role not in ['SUPERADMIN', 'DIRECTOR', 'CHIEF_ENGINEER']):
            return Response(
                {'detail': 'У вас нет прав для восстановления этой позиции'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Проверка, что позиция была отменена
        if item.status != 'CANCELLED':
            return Response(
                {'detail': 'Позицию можно восстановить только если она отменена'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Восстанавливаем статус позиции
        item.status = 'ACTIVE'

        # Восстанавливаем предыдущий статус в процессе согласования
        if item.previous_item_status:
            item.item_status = item.previous_item_status
        else:
            # Если предыдущий статус не сохранен, ставим DRAFT
            item.item_status = 'DRAFT'

        # Очищаем данные об отмене
        item.cancellation_reason = ''
        item.cancelled_by = None
        item.cancelled_at = None
        item.previous_item_status = None

        item.save()

        # Добавляем запись в историю
        from apps.material_requests.models import MaterialRequestHistory
        MaterialRequestHistory.objects.create(
            request=material_request,
            user=request.user,
            old_status=material_request.status,
            new_status=material_request.status,
            comment=f'Позиция "{item.material_name}" восстановлена'
        )

        # Возвращаем обновленную позицию
        serializer = self.get_serializer(item)
        return Response(serializer.data)

    @action(detail=True, methods=['patch'])
    def approve_item(self, request, pk=None):
        """
        Согласование позиции материала директором.
        Endpoint: PATCH /api/material-request-items/{id}/approve_item/

        Body: {
            "approval_status": "APPROVED" | "REJECTED" | "REWORK",
            "comment": "Комментарий (опционально)"
        }
        """
        item = self.get_object()
        material_request = item.request

        # Проверка прав - только директор или суперадмин
        if request.user.role not in ['SUPERADMIN', 'DIRECTOR']:
            return Response(
                {'detail': 'У вас нет прав для согласования этой позиции'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Получаем новый статус согласования
        approval_status = request.data.get('approval_status')
        if not approval_status or approval_status not in ['APPROVED', 'REJECTED', 'REWORK']:
            return Response(
                {'detail': 'Необходимо указать корректный статус согласования'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Обновляем статус согласования позиции
        item.approval_status = approval_status
        item.save()

        # Возвращаем обновленную позицию
        serializer = self.get_serializer(item)
        return Response(serializer.data)

    @action(detail=True, methods=['patch'])
    def update_availability(self, request, pk=None):
        """
        Обновление статуса наличия материала на складе (для зав.центрскладом).
        Endpoint: PATCH /api/material-request-items/{id}/update_availability/

        Body: {
            "availability_status": "IN_STOCK" | "PARTIALLY_IN_STOCK" | "OUT_OF_STOCK",
            "available_quantity": 10.5 (опционально, для PARTIALLY_IN_STOCK)
        }
        """
        item = self.get_object()
        material_request = item.request

        # Проверка прав - только зав.склада или суперадмин
        if request.user.role not in ['SUPERADMIN', 'WAREHOUSE_HEAD']:
            return Response(
                {'detail': 'У вас нет прав для обновления статуса наличия'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Получаем новый статус наличия
        availability_status = request.data.get('availability_status')
        if not availability_status or availability_status not in ['IN_STOCK', 'PARTIALLY_IN_STOCK', 'OUT_OF_STOCK']:
            return Response(
                {'detail': 'Необходимо указать корректный статус наличия'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Обновляем статус наличия позиции
        item.availability_status = availability_status

        # Если частично в наличии, получаем количество
        if availability_status == 'PARTIALLY_IN_STOCK':
            available_quantity = request.data.get('available_quantity')
            if available_quantity is not None:
                item.available_quantity = available_quantity
        else:
            # Для других статусов обнуляем available_quantity
            item.available_quantity = None

        item.save()

        # Возвращаем обновленную позицию
        serializer = self.get_serializer(item)
        return Response(serializer.data)

    @action(detail=True, methods=['patch'])
    def change_item_status(self, request, pk=None):
        """
        Изменение статуса отдельной позиции материала в процессе согласования.
        Endpoint: PATCH /api/material-request-items/{id}/change_item_status/

        Body: {
            "new_status": "UNDER_REVIEW" | "WAREHOUSE_CHECK" | "BACK_TO_SUPPLY" | ...,
            "comment": "Комментарий (опционально)"
        }
        """
        from apps.material_requests.models import MaterialRequestHistory

        item = self.get_object()
        material_request = item.request

        # Получаем новый статус
        new_status = request.data.get('new_status')
        if not new_status:
            return Response(
                {'detail': 'Необходимо указать новый статус'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Проверяем, что новый статус валиден
        valid_statuses = [choice[0] for choice in MaterialRequestItem.ProcessStatus.choices]
        if new_status not in valid_statuses:
            return Response(
                {'detail': f'Некорректный статус. Допустимые значения: {", ".join(valid_statuses)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Проверяем права доступа в зависимости от текущего и нового статуса
        user_role = request.user.role
        current_status = item.item_status

        # Логика проверки прав по новой схеме: Автор → Снабжение → Завсклад → Снабжение → Инженер ПТО → Снабжение → Рук.проекта → Снабжение → Директор → Снабжение
        allowed = False

        # 1. Прораб/мастер/начальник участка может отправлять из DRAFT, REWORK или RETURNED_FOR_REVISION
        if user_role in ['FOREMAN', 'MASTER', 'SITE_MANAGER', 'SUPERADMIN']:
            if current_status in ['DRAFT', 'REWORK', 'RETURNED_FOR_REVISION'] and new_status == 'UNDER_REVIEW':
                allowed = True
            elif current_status in ['DELIVERY', 'WAREHOUSE_SHIPPING'] and new_status == 'COMPLETED':
                allowed = True

        # 2. Снабженец может переводить через множество статусов (новая логика)
        if user_role in ['SUPPLY_MANAGER', 'SUPERADMIN']:
            # Снабжение → Завсклад
            if current_status == 'UNDER_REVIEW' and new_status == 'WAREHOUSE_CHECK':
                allowed = True
            # Снабжение (после склада) → Инженер ПТО
            elif current_status == 'BACK_TO_SUPPLY' and new_status == 'ENGINEER_APPROVAL':
                allowed = True
            # Снабжение (после инженера) → Руководитель проекта
            elif current_status == 'BACK_TO_SUPPLY_AFTER_ENGINEER' and new_status == 'PROJECT_MANAGER_APPROVAL':
                allowed = True
            # Снабжение (после рук.проекта) → Директор
            elif current_status == 'BACK_TO_SUPPLY_AFTER_PM' and new_status == 'DIRECTOR_APPROVAL':
                allowed = True
            # Снабжение (после директора) → Отправка на объект / Оплата
            elif current_status == 'BACK_TO_SUPPLY_AFTER_DIRECTOR' and new_status in ['PAYMENT', 'SENT_TO_SITE']:
                allowed = True
            # Оплата → Оплачено
            elif current_status == 'PAYMENT' and new_status == 'PAID':
                allowed = True
            # Оплачено → Доставлено
            elif current_status == 'PAID' and new_status == 'DELIVERY':
                allowed = True

        # 3. Зав.склада возвращает снабженцу (после склада)
        if user_role in ['WAREHOUSE_HEAD', 'SUPERADMIN']:
            if current_status == 'WAREHOUSE_CHECK' and new_status == 'BACK_TO_SUPPLY':
                allowed = True
            elif current_status == 'SENT_TO_SITE' and new_status == 'WAREHOUSE_SHIPPING':
                allowed = True

        # 4. Инженер ПТО возвращает снабженцу (после инженера) или отправляет на доработку автору
        if user_role in ['ENGINEER', 'SUPERADMIN']:
            if current_status == 'ENGINEER_APPROVAL' and new_status == 'BACK_TO_SUPPLY_AFTER_ENGINEER':
                allowed = True
            # Инженер ПТО может отправить на доработку автору
            elif current_status == 'ENGINEER_APPROVAL' and new_status == 'RETURNED_FOR_REVISION':
                allowed = True

        # 5. Руководитель проекта возвращает снабженцу (после рук.проекта) или отправляет на доработку автору
        if user_role in ['PROJECT_MANAGER', 'SUPERADMIN']:
            if current_status == 'PROJECT_MANAGER_APPROVAL' and new_status == 'BACK_TO_SUPPLY_AFTER_PM':
                allowed = True
            # Руководитель проекта может отправить на доработку автору
            elif current_status == 'PROJECT_MANAGER_APPROVAL' and new_status == 'RETURNED_FOR_REVISION':
                allowed = True

        # 6. Директор и Главный инженер возвращают снабженцу (после директора) или отправляют на доработку автору
        if user_role in ['DIRECTOR', 'CHIEF_ENGINEER', 'SUPERADMIN']:
            if current_status == 'DIRECTOR_APPROVAL' and new_status == 'BACK_TO_SUPPLY_AFTER_DIRECTOR':
                allowed = True
            # Директор и Главный инженер могут отправить на доработку автору
            elif current_status == 'DIRECTOR_APPROVAL' and new_status == 'RETURNED_FOR_REVISION':
                allowed = True

        if not allowed:
            return Response(
                {'detail': 'У вас нет прав для изменения статуса этой позиции'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Сохраняем старый статус для истории
        old_status = item.item_status

        # Обновляем статус позиции
        item.item_status = new_status
        item.save()

        # Создаем запись в истории (можно добавить comment из request.data)
        comment = request.data.get('comment', f'Статус позиции изменен: {old_status} → {new_status}')
        MaterialRequestHistory.objects.create(
            request=material_request,
            user=request.user,
            old_status=old_status,
            new_status=new_status,
            comment=f'Позиция "{item.material_name}": {comment}'
        )

        # Возвращаем обновленную позицию
        serializer = self.get_serializer(item)
        return Response(serializer.data)

    @action(detail=True, methods=['patch'])
    def record_actual_quantity(self, request, pk=None):
        """
        Фиксация фактического количества поступившего материала.
        Создает запись на складе и обновляет actual_quantity.

        Endpoint: PATCH /api/material-request-items/{id}/record_actual_quantity/

        Body: {
            "actual_quantity": 100.50,
            "receipt_date": "2025-10-22T10:30:00Z" (опционально, по умолчанию текущее время),
            "waybill_number": "ТТН-12345" (опционально),
            "supplier": "ООО Поставщик" (опционально),
            "quality_status": "GOOD" | "DAMAGED" | "DEFECTIVE" | "PARTIAL" (опционально, по умолчанию GOOD),
            "notes": "Примечание" (опционально)
        }
        """
        from apps.warehouse.models import WarehouseReceipt
        from datetime import datetime

        item = self.get_object()
        material_request = item.request
        user = request.user

        # Проверяем права: только автор может фиксировать фактическое количество
        if user.role != 'SUPERADMIN' and material_request.author != user:
            return Response(
                {'detail': 'Только автор заявки может фиксировать фактическое количество'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Получаем actual_quantity из запроса
        actual_quantity = request.data.get('actual_quantity')
        if actual_quantity is None:
            return Response(
                {'detail': 'Необходимо указать фактическое количество (actual_quantity)'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            actual_quantity = float(actual_quantity)
            if actual_quantity < 0:
                return Response(
                    {'detail': 'Количество должно быть положительным числом'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except (ValueError, TypeError):
            return Response(
                {'detail': 'Некорректное значение для actual_quantity'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Получаем дополнительные параметры
        receipt_date = request.data.get('receipt_date')
        if receipt_date:
            try:
                receipt_date = datetime.fromisoformat(receipt_date.replace('Z', '+00:00'))
            except ValueError:
                return Response(
                    {'detail': 'Некорректный формат даты'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        else:
            receipt_date = datetime.now()

        waybill_number = request.data.get('waybill_number', '')
        supplier = request.data.get('supplier', '')
        quality_status = request.data.get('quality_status', 'GOOD')
        notes = request.data.get('notes', '')

        # Создаем запись на складе
        try:
            warehouse_receipt = WarehouseReceipt.objects.create(
                material_request=material_request,
                material_item=item,
                project=material_request.project,
                receipt_date=receipt_date,
                received_quantity=actual_quantity,
                unit=item.unit,
                waybill_number=waybill_number,
                supplier=supplier,
                received_by=user,
                quality_status=quality_status,
                notes=notes
            )
        except Exception as e:
            return Response(
                {'detail': f'Ошибка при создании записи на складе: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        # Обновляем actual_quantity в позиции материала
        item.actual_quantity = actual_quantity
        item.save(update_fields=['actual_quantity'])

        # Возвращаем обновленную позицию
        serializer = self.get_serializer(item)
        return Response({
            'item': serializer.data,
            'warehouse_receipt_id': warehouse_receipt.id,
            'message': 'Поступление зафиксировано на складе'
        })


class MaterialRequestDocumentViewSet(viewsets.ModelViewSet):
    """ViewSet для работы с документами заявок."""

    queryset = MaterialRequestDocument.objects.all().select_related('request', 'uploaded_by')
    serializer_class = MaterialRequestDocumentSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['request', 'document_type']


class MaterialRequestCommentViewSet(viewsets.ModelViewSet):
    """ViewSet для работы с комментариями к заявкам."""

    queryset = MaterialRequestComment.objects.all().select_related('request', 'author')
    serializer_class = MaterialRequestCommentSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['request', 'author']
