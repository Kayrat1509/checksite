"""
Обработчик Excel для импорта/экспорта надзоров (Технадзор и Авторский надзор).

Поддерживает три варианта работы:
1. Массовое добавление (создание новых надзоров)
2. Обновление данных (export → edit → import)
3. Backup (полная выгрузка с timestamp)

Особенности:
- Генерация постоянных паролей (не временных)
- Валидация email, телефонов, проектов
- Dropdown списки для ролей (SUPERVISOR, OBSERVER) и проектов
- Отправка credentials на email
- Роли: SUPERVISOR (Технадзор), OBSERVER (Авторский надзор)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import re
import logging

logger = logging.getLogger(__name__)


class SupervisionExcelHandler:
    """
    Класс для работы с Excel файлами надзоров.

    Методы:
    - generate_template_v2() — создать шаблон для импорта с dropdown
    - generate_export_v2() — экспорт текущих надзоров
    - generate_backup() — создать backup с timestamp
    - parse_import_file() — парсинг и валидация импортируемого файла
    """

    # Константы для Excel
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # Синий для надзоров
    HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    CELL_BORDER = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Заголовки колонок для надзоров
    COLUMNS = {
        'email': 'Email*',
        'full_name': 'ФИО*',
        'role': 'Роль*',
        'phone': 'Телефон',
        'projects': 'Объекты'
    }

    # Ширина колонок (в символах)
    COLUMN_WIDTHS = {
        'A': 30,  # Email
        'B': 30,  # ФИО
        'C': 25,  # Роль
        'D': 18,  # Телефон
        'E': 40   # Объекты
    }

    # Роли надзоров
    ROLES = {
        'SUPERVISOR': 'Технадзор',
        'OBSERVER': 'Авторский надзор'
    }

    def __init__(self, company):
        """
        Args:
            company: Company instance для изоляции данных
        """
        self.company = company

    def generate_template_v2(self):
        """
        Генерация шаблона Excel v2 для импорта надзоров.

        Структура:
        - Лист 1: "Инструкция" (ReadOnly) — описание процесса импорта
        - Лист 2: "Данные" (Editable) — таблица с данными и dropdown

        Returns:
            openpyxl.Workbook: Готовый workbook для сохранения
        """
        workbook = openpyxl.Workbook()

        # ===== ЛИСТ 1: ИНСТРУКЦИЯ =====
        ws_instruction = workbook.active
        ws_instruction.title = "Инструкция"

        instructions = [
            ("ИНСТРУКЦИЯ ПО ИМПОРТУ НАДЗОРОВ", 16, True),
            ("", 11, False),
            ("1. ОБЯЗАТЕЛЬНЫЕ ПОЛЯ (отмечены звездочкой *):", 11, True),
            ("   - Email* — уникальный email надзора", 11, False),
            ("   - ФИО* — полное имя (Фамилия Имя Отчество или Фамилия Имя)", 11, False),
            ("   - Роль* — Технадзор или Авторский надзор", 11, False),
            ("", 11, False),
            ("2. НЕОБЯЗАТЕЛЬНЫЕ ПОЛЯ:", 11, True),
            ("   - Телефон — контактный телефон (любой формат)", 11, False),
            ("   - Объекты — можно выбрать из выпадающего списка ИЛИ ввести несколько через запятую:", 11, False),
            ("     Пример 1: Выбрать из списка → Жилой комплекс Восход", 11, False),
            ("     Пример 2: Ввести несколько → Жилой комплекс Восход, ТЦ Запад, Школа №5", 11, False),
            ("     (при вводе нескольких объектов разделяйте их запятой)", 11, False),
            ("", 11, False),
            ("3. ВАЖНО:", 11, True),
            ("   - Email должен быть уникальным в системе", 11, False),
            ("   - При импорте будет сгенерирован постоянный пароль (12 символов)", 11, False),
            ("   - Данные для входа будут отправлены на указанный email", 11, False),
            ("   - Надзоры будут привязаны к вашей компании автоматически", 11, False),
            ("   - Рекомендуется сменить пароль в профиле после первого входа", 11, False),
            ("", 11, False),
            ("4. ПОРЯДОК РАБОТЫ:", 11, True),
            ("   1) Заполните таблицу на листе 'Данные'", 11, False),
            ("   2) Сохраните файл", 11, False),
            ("   3) Загрузите через кнопку 'Импорт' на странице Надзоры", 11, False),
            ("   4) Дождитесь завершения импорта и проверьте статистику", 11, False),
            ("", 11, False),
            ("5. ПРИМЕР ЗАПОЛНЕНИЯ:", 11, True),
            ("   Email: ivanov@supervision.kz", 11, False),
            ("   ФИО: Иванов Иван Иванович", 11, False),
            ("   Роль: Технадзор", 11, False),
            ("   Телефон: +7 777 123 45 67", 11, False),
            ("   Объекты: Жилой комплекс \"Север\", ТЦ \"Запад\"", 11, False),
            ("", 11, False),
            (f"Дата создания шаблона: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 10, False),
            (f"Компания: {self.company.name}", 10, False),
        ]

        # Заполняем инструкцию
        for idx, (text, font_size, is_bold) in enumerate(instructions, start=1):
            cell = ws_instruction.cell(row=idx, column=1, value=text)
            cell.font = Font(name='Arial', size=font_size, bold=is_bold)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Устанавливаем ширину колонки A для инструкции
        ws_instruction.column_dimensions['A'].width = 100

        # ===== ЛИСТ 2: ДАННЫЕ =====
        ws_data = workbook.create_sheet("Данные")

        # Заголовки
        headers = list(self.COLUMNS.values())
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.CELL_BORDER

        # Устанавливаем ширину колонок
        for col_letter, width in self.COLUMN_WIDTHS.items():
            ws_data.column_dimensions[col_letter].width = width

        # Устанавливаем высоту заголовка
        ws_data.row_dimensions[1].height = 30

        # ===== DROPDOWN ДЛЯ РОЛЕЙ =====
        roles_list = ','.join(self.ROLES.values())
        roles_validation = DataValidation(
            type="list",
            formula1=f'"{roles_list}"',
            allow_blank=False,
            showErrorMessage=True,
            errorTitle='Неверная роль',
            error='Выберите роль из списка: Технадзор или Авторский надзор',
            showInputMessage=True,
            promptTitle='Выбор роли',
            prompt='Выберите роль из списка'
        )
        roles_validation.add('C2:C1000')  # Применяем к колонке "Роль" (C)
        ws_data.add_data_validation(roles_validation)

        # ===== DROPDOWN ДЛЯ ПРОЕКТОВ/ОБЪЕКТОВ =====
        company_projects = self.company.projects.filter(is_active=True).values_list('name', flat=True)

        if company_projects:
            project_list = ','.join(company_projects)

            if len(project_list) <= 250:
                # Если список проектов короткий, используем прямую формулу
                projects_formula = f'"{project_list}"'

                projects_validation = DataValidation(
                    type="list",
                    formula1=projects_formula,
                    allow_blank=True,
                    showErrorMessage=False,
                    showInputMessage=True,
                    promptTitle='Выбор объектов',
                    prompt='Выберите один объект из списка или введите несколько через запятую'
                )
                projects_validation.add('E2:E1000')  # Применяем к колонке "Объекты" (E)
                ws_data.add_data_validation(projects_validation)
            else:
                # Если список длинный, создаем скрытый лист с проектами
                ws_projects = workbook.create_sheet("__Объекты__")

                for idx, project_name in enumerate(company_projects, start=1):
                    ws_projects.cell(row=idx, column=1, value=project_name)

                # Скрываем лист
                ws_projects.sheet_state = 'hidden'

                # Создаем валидацию с reference на скрытый лист
                projects_validation = DataValidation(
                    type="list",
                    formula1=f"'__Объекты__'!$A$1:$A${len(company_projects)}",
                    allow_blank=True,
                    showErrorMessage=False,
                    showInputMessage=True,
                    promptTitle='Выбор объектов',
                    prompt='Выберите один объект из списка или введите несколько через запятую'
                )
                projects_validation.add('E2:E1000')
                ws_data.add_data_validation(projects_validation)

        # Добавляем примеры данных (2 строки)
        examples = [
            {
                'email': 'technadz or@company.kz',
                'full_name': 'Иванов Иван Иванович',
                'role': 'Технадзор',
                'phone': '+7 777 123 45 67',
                'projects': '' if not company_projects else list(company_projects)[0]
            },
            {
                'email': 'observer@company.kz',
                'full_name': 'Петров Петр Петрович',
                'role': 'Авторский надзор',
                'phone': '+7 777 987 65 43',
                'projects': '' if not company_projects else ', '.join(list(company_projects)[:2]) if len(company_projects) >= 2 else list(company_projects)[0]
            }
        ]

        for row_idx, example in enumerate(examples, start=2):
            for col_idx, field in enumerate(self.COLUMNS.keys(), start=1):
                cell = ws_data.cell(row=row_idx, column=col_idx, value=example.get(field, ''))
                cell.border = self.CELL_BORDER
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        return workbook

    def generate_export_v2(self):
        """
        Экспорт текущих надзоров компании в Excel v2.

        Создает файл с текущими данными для редактирования и последующего
        реимпорта в режиме update.

        Returns:
            openpyxl.Workbook: Готовый workbook для сохранения
        """
        from apps.users.models import User

        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = "Надзоры"

        # Заголовки
        headers = list(self.COLUMNS.values())
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.CELL_BORDER

        # Устанавливаем ширину колонок
        for col_letter, width in self.COLUMN_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width

        # Устанавливаем высоту заголовка
        ws.row_dimensions[1].height = 30

        # ===== DROPDOWN ДЛЯ РОЛЕЙ =====
        roles_list = ','.join(self.ROLES.values())
        roles_validation = DataValidation(
            type="list",
            formula1=f'"{roles_list}"',
            allow_blank=False,
            showErrorMessage=True,
            errorTitle='Неверная роль',
            error='Выберите роль из списка'
        )
        roles_validation.add('C2:C1000')
        ws.add_data_validation(roles_validation)

        # ===== DROPDOWN ДЛЯ ПРОЕКТОВ =====
        company_projects = self.company.projects.filter(is_active=True).values_list('name', flat=True)

        if company_projects:
            project_list = ','.join(company_projects)

            if len(project_list) <= 250:
                projects_formula = f'"{project_list}"'

                projects_validation = DataValidation(
                    type="list",
                    formula1=projects_formula,
                    allow_blank=True,
                    showErrorMessage=False,
                    showInputMessage=True,
                    promptTitle='Выбор объектов',
                    prompt='Выберите один объект из списка или введите несколько через запятую'
                )
                projects_validation.add('E2:E1000')
                ws.add_data_validation(projects_validation)
            else:
                ws_projects = workbook.create_sheet("__Объекты__")

                for idx, project_name in enumerate(company_projects, start=1):
                    ws_projects.cell(row=idx, column=1, value=project_name)

                ws_projects.sheet_state = 'hidden'

                projects_validation = DataValidation(
                    type="list",
                    formula1=f"'__Объекты__'!$A$1:$A${len(company_projects)}",
                    allow_blank=True,
                    showErrorMessage=False,
                    showInputMessage=True,
                    promptTitle='Выбор объектов',
                    prompt='Выберите один объект из списка или введите несколько через запятую'
                )
                projects_validation.add('E2:E1000')
                ws.add_data_validation(projects_validation)

        # ===== ДАННЫЕ НАДЗОРОВ =====
        # Получаем только надзоров компании (SUPERVISOR, OBSERVER), не удаленных
        supervisions = User.objects.filter(
            company=self.company,
            role__in=['SUPERVISOR', 'OBSERVER'],
            is_deleted=False
        ).prefetch_related('projects')

        for row_idx, supervision in enumerate(supervisions, start=2):
            # Формируем список проектов через запятую
            project_names = ', '.join([p.name for p in supervision.projects.all()])

            # Формируем ФИО (Фамилия Имя Отчество)
            full_name_parts = [supervision.last_name, supervision.first_name]
            if supervision.middle_name:
                full_name_parts.append(supervision.middle_name)
            full_name = ' '.join(full_name_parts)

            # Получаем русское название роли
            role_display = self.ROLES.get(supervision.role, supervision.role)

            row_data = {
                'email': supervision.email,
                'full_name': full_name,
                'role': role_display,
                'phone': supervision.phone or '',
                'projects': project_names
            }

            for col_idx, field in enumerate(self.COLUMNS.keys(), start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(field, ''))
                cell.border = self.CELL_BORDER
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        return workbook

    def generate_backup(self):
        """
        Создание полного backup надзоров с timestamp.

        Включает все данные, включая статусы и даты для архивных целей.

        Returns:
            openpyxl.Workbook: Готовый workbook для сохранения
        """
        from apps.users.models import User

        workbook = openpyxl.Workbook()

        # ===== ЛИСТ 1: ИНФОРМАЦИЯ О BACKUP =====
        ws_info = workbook.active
        ws_info.title = "Информация"

        backup_info = [
            ("BACKUP НАДЗОРОВ", 16, True),
            ("", 11, False),
            (f"Дата создания: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 11, False),
            (f"Компания: {self.company.name}", 11, False),
            ("", 11, False),
            ("Содержание:", 11, True),
            ("- Лист 'Надзоры': все надзоры компании (активные и архивированные)", 11, False),
            ("- Включает расширенные данные: ID, статусы, даты", 11, False),
            ("", 11, False),
            ("Назначение:", 11, True),
            ("- Архивное хранение данных", 11, False),
            ("- Восстановление в случае потери данных", 11, False),
            ("- Анализ и отчетность", 11, False),
        ]

        for idx, (text, font_size, is_bold) in enumerate(backup_info, start=1):
            cell = ws_info.cell(row=idx, column=1, value=text)
            cell.font = Font(name='Arial', size=font_size, bold=is_bold)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

        ws_info.column_dimensions['A'].width = 80

        # ===== ЛИСТ 2: ДАННЫЕ НАДЗОРОВ =====
        ws_data = workbook.create_sheet("Надзоры")

        # Расширенные заголовки для backup
        backup_columns = {
            'id': 'ID',
            'email': 'Email',
            'full_name': 'ФИО',
            'role': 'Роль',
            'phone': 'Телефон',
            'projects': 'Объекты',
            'is_active': 'Активен',
            'is_deleted': 'В корзине',
            'created_at': 'Дата создания',
            'updated_at': 'Дата обновления'
        }

        # Заголовки
        for col_idx, header in enumerate(backup_columns.values(), start=1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.CELL_BORDER

        # Устанавливаем ширину колонок
        column_widths = [6, 30, 30, 25, 18, 40, 10, 12, 20, 20]
        for idx, width in enumerate(column_widths, start=1):
            ws_data.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

        # Устанавливаем высоту заголовка
        ws_data.row_dimensions[1].height = 30

        # ===== ДАННЫЕ НАДЗОРОВ (ВСЕ, включая архивированных) =====
        supervisions = User.objects.filter(
            company=self.company,
            role__in=['SUPERVISOR', 'OBSERVER']
        ).prefetch_related('projects')

        for row_idx, supervision in enumerate(supervisions, start=2):
            # Формируем список проектов через запятую
            project_names = ', '.join([p.name for p in supervision.projects.all()])

            # Формируем ФИО (Фамилия Имя Отчество)
            full_name_parts = [supervision.last_name, supervision.first_name]
            if supervision.middle_name:
                full_name_parts.append(supervision.middle_name)
            full_name = ' '.join(full_name_parts)

            # Получаем русское название роли
            role_display = self.ROLES.get(supervision.role, supervision.role)

            row_data = {
                'id': supervision.id,
                'email': supervision.email,
                'full_name': full_name,
                'role': role_display,
                'phone': supervision.phone or '',
                'projects': project_names,
                'is_active': 'Да' if supervision.is_active else 'Нет',
                'is_deleted': 'Да' if supervision.is_deleted else 'Нет',
                'created_at': supervision.created_at.strftime('%Y-%m-%d %H:%M:%S') if supervision.created_at else '',
                'updated_at': supervision.updated_at.strftime('%Y-%m-%d %H:%M:%S') if supervision.updated_at else ''
            }

            for col_idx, field in enumerate(backup_columns.keys(), start=1):
                cell = ws_data.cell(row=row_idx, column=col_idx, value=row_data.get(field, ''))
                cell.border = self.CELL_BORDER
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        return workbook

    def parse_import_file(self, file, mode='create'):
        """
        Парсинг и валидация Excel файла для импорта надзоров.

        Args:
            file: Upload file object (Excel файл)
            mode: 'create' или 'update'

        Returns:
            dict: {
                'valid_rows': [список валидных строк],
                'errors': [список ошибок валидации]
            }
        """
        from apps.users.models import User
        from apps.projects.models import Project

        try:
            workbook = openpyxl.load_workbook(file)
            ws = workbook['Данные'] if 'Данные' in workbook.sheetnames else workbook.active
        except Exception as e:
            logger.error(f"Ошибка при чтении Excel файла: {str(e)}")
            return {
                'valid_rows': [],
                'errors': [{'row': 0, 'errors': [f'Не удалось прочитать файл: {str(e)}']}]
            }

        valid_rows = []
        errors = []

        # Обратный маппинг: русское название → код роли
        role_reverse_map = {v: k for k, v in self.ROLES.items()}

        # Начинаем с 2-й строки (1-я строка — заголовки)
        for row_idx in range(2, ws.max_row + 1):
            row_errors = []

            # Читаем данные из ячеек (Email, ФИО, Роль, Телефон, Объекты)
            email = ws.cell(row=row_idx, column=1).value
            full_name = ws.cell(row=row_idx, column=2).value
            role_display = ws.cell(row=row_idx, column=3).value
            phone = ws.cell(row=row_idx, column=4).value
            projects_str = ws.cell(row=row_idx, column=5).value

            # Пропускаем пустые строки
            if not any([email, full_name, role_display]):
                continue

            # === ВАЛИДАЦИЯ ОБЯЗАТЕЛЬНЫХ ПОЛЕЙ ===

            # Email
            if not email or not isinstance(email, str):
                row_errors.append('Email обязателен')
            else:
                email = email.strip().lower()
                # Проверка формата email
                if not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
                    row_errors.append('Некорректный формат email')
                else:
                    # Проверка уникальности email
                    if mode == 'create':
                        if User.objects.filter(email__iexact=email).exists():
                            row_errors.append(f'Email {email} уже существует в системе')
                    elif mode == 'update':
                        if not User.objects.filter(
                            email__iexact=email,
                            company=self.company,
                            role__in=['SUPERVISOR', 'OBSERVER']
                        ).exists():
                            row_errors.append(f'Надзор с email {email} не найден')

            # ФИО (парсим на фамилию, имя, отчество)
            if not full_name or not isinstance(full_name, str) or not full_name.strip():
                row_errors.append('ФИО обязательно')
                last_name = ''
                first_name = ''
                middle_name = ''
            else:
                full_name = full_name.strip()
                name_parts = full_name.split()

                if len(name_parts) < 2:
                    row_errors.append('ФИО должно содержать минимум Фамилию и Имя')
                    last_name = name_parts[0] if len(name_parts) > 0 else ''
                    first_name = ''
                    middle_name = ''
                else:
                    last_name = name_parts[0]
                    first_name = name_parts[1]
                    middle_name = ' '.join(name_parts[2:]) if len(name_parts) > 2 else ''

            # Роль
            if not role_display or not isinstance(role_display, str):
                row_errors.append('Роль обязательна')
                role = None
            else:
                role_display = role_display.strip()
                role = role_reverse_map.get(role_display)
                if not role:
                    row_errors.append(f'Неверная роль: {role_display}. Используйте: Технадзор или Авторский надзор')

            # === ВАЛИДАЦИЯ НЕОБЯЗАТЕЛЬНЫХ ПОЛЕЙ ===

            # Телефон (необязательное поле, любой формат разрешён)
            if phone:
                phone = str(phone).strip()
            else:
                phone = ''

            # Проекты/Объекты
            project_ids = []
            if projects_str:
                projects_str = str(projects_str).strip()
                # Разбиваем по запятой
                project_names = [p.strip() for p in projects_str.split(',') if p.strip()]

                for project_name in project_names:
                    try:
                        project = Project.objects.get(
                            name__iexact=project_name,
                            company=self.company,
                            is_active=True
                        )
                        project_ids.append(project.id)
                    except Project.DoesNotExist:
                        row_errors.append(f'Проект "{project_name}" не найден')
                    except Project.MultipleObjectsReturned:
                        row_errors.append(f'Найдено несколько проектов с названием "{project_name}"')

            # Если есть ошибки валидации, добавляем в список ошибок
            if row_errors:
                errors.append({
                    'row': row_idx,
                    'errors': row_errors
                })
            else:
                # Добавляем валидную строку
                valid_rows.append({
                    'email': email,
                    'last_name': last_name,
                    'first_name': first_name,
                    'middle_name': middle_name,
                    'role': role,
                    'phone': phone,
                    'project_ids': project_ids
                })

        return {
            'valid_rows': valid_rows,
            'errors': errors
        }
