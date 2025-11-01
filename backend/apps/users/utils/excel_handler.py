"""
Обработчик Excel для импорта/экспорта персонала.

Поддерживает три варианта работы:
1. Массовое добавление (создание новых сотрудников)
2. Обновление данных (export → edit → import)
3. Backup (полная выгрузка с timestamp)

Особенности:
- Генерация постоянных паролей (не временных)
- Валидация email, телефонов, проектов
- Dropdown списки для ролей и проектов
- Отправка credentials на email
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import re
import logging

logger = logging.getLogger(__name__)


class PersonnelExcelHandler:
    """
    Класс для работы с Excel файлами персонала.

    Методы:
    - generate_template_v2() — создать шаблон для импорта с dropdown
    - generate_export_v2() — экспорт текущих пользователей
    - generate_backup() — создать backup с timestamp
    - parse_import_file() — парсинг и валидация импортируемого файла
    """

    # Константы для Excel
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    CELL_BORDER = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Заголовки колонок
    COLUMNS = {
        'email': 'Email*',
        'full_name': 'ФИО*',
        'role': 'Роль*',
        'position': 'Должность',
        'phone': 'Телефон',
        'projects': 'Проекты'
    }

    # Ширина колонок (в символах)
    COLUMN_WIDTHS = {
        'A': 30,  # Email
        'B': 25,  # ФИО
        'C': 25,  # Роль
        'D': 20,  # Должность
        'E': 15,  # Телефон
        'F': 40   # Проекты
    }

    def __init__(self, company):
        """
        Args:
            company: Company instance для изоляции данных
        """
        self.company = company

    def generate_template_v2(self):
        """
        Генерация шаблона Excel v2 для импорта персонала.

        Структура:
        - Лист 1: "Инструкция" (ReadOnly) — описание процесса импорта
        - Лист 2: "Данные" (Editable) — таблица с данными и dropdown

        Returns:
            openpyxl.Workbook: Готовый workbook для сохранения
        """
        from apps.users.models import User

        workbook = openpyxl.Workbook()

        # ===== ЛИСТ 1: ИНСТРУКЦИЯ =====
        ws_instruction = workbook.active
        ws_instruction.title = "Инструкция"

        instructions = [
            ("ИНСТРУКЦИЯ ПО ИМПОРТУ ПЕРСОНАЛА", 16, True),
            ("", 11, False),
            ("1. ОБЯЗАТЕЛЬНЫЕ ПОЛЯ (отмечены звездочкой *):", 11, True),
            ("   - Email* — уникальный email сотрудника", 11, False),
            ("   - ФИО* — полное имя (Фамилия Имя Отчество)", 11, False),
            ("   - Роль* — выберите из выпадающего списка", 11, False),
            ("", 11, False),
            ("2. НЕОБЯЗАТЕЛЬНЫЕ ПОЛЯ:", 11, True),
            ("   - Должность — текстовое описание должности", 11, False),
            ("   - Телефон — контактный телефон (любой формат)", 11, False),
            ("   - Проекты — можно выбрать из выпадающего списка ИЛИ ввести несколько через запятую:", 11, False),
            ("     Пример 1: Выбрать из списка → Жилой комплекс Восход", 11, False),
            ("     Пример 2: Ввести несколько → Жилой комплекс Восход, ТЦ Запад, Школа №5", 11, False),
            ("     (при вводе нескольких проектов разделяйте их запятой)", 11, False),
            ("", 11, False),
            ("3. ВАЖНО:", 11, True),
            ("   - Email должен быть уникальным в системе", 11, False),
            ("   - При импорте будет сгенерирован постоянный пароль (12 символов)", 11, False),
            ("   - Данные для входа будут отправлены на указанный email", 11, False),
            ("   - Пользователи будут привязаны к вашей компании автоматически", 11, False),
            ("   - Рекомендуется сменить пароль в профиле после первого входа", 11, False),
            ("", 11, False),
            ("4. ОПИСАНИЕ РОЛЕЙ:", 11, True),
            ("", 11, False),
            ("   КАТЕГОРИЯ \"РУКОВОДСТВО\" (полный доступ ко всем страницам):", 11, True),
            ("   - Директор (полный доступ к системе)", 11, False),
            ("   - Главный инженер (полный доступ к системе)", 11, False),
            ("   - Руководитель проекта (управление проектами)", 11, False),
            ("   - Начальник участка (управление участком)", 11, False),
            ("   - Прораб (контроль работ)", 11, False),
            ("", 11, False),
            ("   КАТЕГОРИЯ \"ИТР И СНАБЖЕНИЕ\" (ограниченный доступ):", 11, True),
            ("   - Инженер ПТО (проектно-техническая документация)", 11, False),
            ("   - Мастер (контроль работ на участке)", 11, False),
            ("   - Снабженец (заявки на материалы)", 11, False),
            ("   - Зав.Центрсклада (управление центральным складом)", 11, False),
            ("   - Завсклад объекта (управление складом объекта)", 11, False),
            ("", 11, False),
            ("   ПРОЧИЕ РОЛИ (специальный доступ):", 11, True),
            ("   - Подрядчик (управляется на отдельной странице)", 11, False),
            ("   - Технадзор (управляется на отдельной странице)", 11, False),
            ("   - Наблюдатель (только просмотр)", 11, False),
            ("", 11, False),
            ("5. ФОРМАТ ДАННЫХ:", 11, True),
            ("   - Email: example@company.com", 11, False),
            ("   - ФИО: Иванов Иван Иванович", 11, False),
            ("   - Роль: Директор (выберите из выпадающего списка)", 11, False),
            ("   - Должность: Генеральный директор", 11, False),
            ("   - Телефон: +7 777 123 45 67 (любой формат)", 11, False),
            ("   - Проекты: Жилой комплекс \"Север\", ТЦ \"Запад\" (несколько через запятую)", 11, False),
            ("", 11, False),
            (f"Дата создания шаблона: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 10, False),
            (f"Компания: {self.company.name}", 10, False),
        ]

        # Заполняем инструкцию
        for idx, (text, font_size, is_bold) in enumerate(instructions, start=1):
            cell = ws_instruction.cell(row=idx, column=1, value=text)
            cell.font = Font(name='Arial', size=font_size, bold=is_bold)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Устанавливаем ширину колонки A для инструкции
        ws_instruction.column_dimensions['A'].width = 100

        # ===== ЛИСТ 2: ДАННЫЕ =====
        ws_data = workbook.create_sheet("Данные")

        # Заголовки
        headers = list(self.COLUMNS.values())
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.CELL_BORDER

        # Устанавливаем ширину колонок
        for col_letter, width in self.COLUMN_WIDTHS.items():
            ws_data.column_dimensions[col_letter].width = width

        # Устанавливаем высоту заголовка
        ws_data.row_dimensions[1].height = 30

        # ===== DROPDOWN ДЛЯ РОЛЕЙ =====
        # Все роли из User.Role (кроме SUPERADMIN)
        # Используем role[1] для получения русских названий вместо кодов
        # str() нужен для конвертации Django lazy translation (__proxy__) в обычную строку
        role_choices = [str(role[1]) for role in User.Role.choices if role[0] != 'SUPERADMIN']
        role_formula = f'"{",".join(role_choices)}"'

        role_validation = DataValidation(
            type="list",
            formula1=role_formula,
            allow_blank=False,
            showErrorMessage=True,
            error='Выберите роль из списка',
            errorTitle='Неверная роль'
        )
        role_validation.add('C2:C1000')  # Применяем к колонке "Роль" (C)
        ws_data.add_data_validation(role_validation)

        # ===== DROPDOWN ДЛЯ ПРОЕКТОВ =====
        # Получаем проекты компании
        company_projects = self.company.projects.filter(is_active=True).values_list('name', flat=True)

        if company_projects:
            # Ограничение Excel: формула не может быть длиннее 255 символов
            project_list = ','.join(company_projects)

            if len(project_list) <= 250:
                # Если список проектов короткий, используем формулу
                projects_formula = f'"{project_list}"'

                projects_validation = DataValidation(
                    type="list",
                    formula1=projects_formula,
                    allow_blank=True,
                    showErrorMessage=False,  # Разрешаем свободный ввод для множественного выбора
                    showInputMessage=True,
                    promptTitle='Выбор проектов',
                    prompt='Выберите один проект из списка или введите несколько через запятую (например: Проект 1, Проект 2)'
                )
                projects_validation.add('F2:F1000')  # Применяем к колонке "Проекты" (F)
                ws_data.add_data_validation(projects_validation)
            else:
                # Если список длинный, создаем скрытый лист с проектами
                ws_projects = workbook.create_sheet("__Проекты__")

                for idx, project_name in enumerate(company_projects, start=1):
                    ws_projects.cell(row=idx, column=1, value=project_name)

                # Скрываем лист
                ws_projects.sheet_state = 'hidden'

                # Создаем валидацию с reference на скрытый лист
                projects_validation = DataValidation(
                    type="list",
                    formula1=f"'__Проекты__'!$A$1:$A${len(company_projects)}",
                    allow_blank=True,
                    showErrorMessage=False,  # Разрешаем свободный ввод для множественного выбора
                    showInputMessage=True,
                    promptTitle='Выбор проектов',
                    prompt='Выберите один проект из списка или введите несколько через запятую'
                )
                projects_validation.add('F2:F1000')
                ws_data.add_data_validation(projects_validation)

        # ===== ПРИМЕР СТРОКИ =====
        # Используем русское название роли вместо кода
        example_data = [
            'example@company.com',
            'Иванов Иван Иванович',
            'Директор',  # Русское название вместо 'DIRECTOR'
            'Генеральный директор',
            '+79991234567',
            ', '.join(list(company_projects)[:2]) if company_projects else ''
        ]

        for col_idx, value in enumerate(example_data, start=1):
            cell = ws_data.cell(row=2, column=col_idx, value=value)
            cell.font = Font(name='Arial', size=10, italic=True, color='808080')
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = self.CELL_BORDER

        # Замораживаем первую строку (заголовок)
        ws_data.freeze_panes = 'A2'

        return workbook

    def generate_export_v2(self):
        """
        Генерация Excel с текущими пользователями компании.

        Аналогична шаблону, но с реальными данными пользователей.

        Returns:
            openpyxl.Workbook: Готовый workbook для сохранения
        """
        from apps.users.models import User

        workbook = openpyxl.Workbook()
        ws_data = workbook.active
        ws_data.title = "Данные"

        # Заголовки
        headers = list(self.COLUMNS.values())
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.CELL_BORDER

        # Устанавливаем ширину колонок
        for col_letter, width in self.COLUMN_WIDTHS.items():
            ws_data.column_dimensions[col_letter].width = width

        ws_data.row_dimensions[1].height = 30

        # Получаем пользователей компании (исключая суперадминов)
        users = User.objects.filter(
            company=self.company
        ).exclude(
            is_superuser=True
        ).select_related('company').prefetch_related('projects').order_by('email')

        # Заполняем данными
        for row_idx, user in enumerate(users, start=2):
            # ФИО
            full_name = f"{user.last_name} {user.first_name}"
            if user.middle_name:
                full_name += f" {user.middle_name}"

            # Проекты
            user_projects = user.projects.filter(is_active=True).values_list('name', flat=True)
            projects_str = ', '.join(user_projects) if user_projects else ''

            # Данные строки
            # Используем get_role_display() для получения русского названия роли
            row_data = [
                user.email,
                full_name.strip(),
                user.get_role_display(),  # Русское название роли вместо кода
                user.position or '',
                user.phone or '',
                projects_str
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name='Arial', size=10)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = self.CELL_BORDER

        # Добавляем dropdown валидацию для ролей и проектов (как в шаблоне)
        # Используем role[1] для получения русских названий вместо кодов
        # str() нужен для конвертации Django lazy translation (__proxy__) в обычную строку
        role_choices = [str(role[1]) for role in User.Role.choices if role[0] != 'SUPERADMIN']
        role_formula = f'"{",".join(role_choices)}"'

        role_validation = DataValidation(
            type="list",
            formula1=role_formula,
            allow_blank=False
        )
        role_validation.add(f'C2:C{len(users) + 1}')
        ws_data.add_data_validation(role_validation)

        # Dropdown для проектов (разрешаем множественный выбор)
        company_projects = self.company.projects.filter(is_active=True).values_list('name', flat=True)
        if company_projects:
            project_list = ','.join(company_projects)
            if len(project_list) <= 250:
                projects_formula = f'"{project_list}"'
                projects_validation = DataValidation(
                    type="list",
                    formula1=projects_formula,
                    allow_blank=True,
                    showErrorMessage=False,  # Разрешаем свободный ввод для множественного выбора
                    showInputMessage=True,
                    promptTitle='Выбор проектов',
                    prompt='Выберите один проект из списка или введите несколько через запятую'
                )
                projects_validation.add(f'F2:F{len(users) + 1}')
                ws_data.add_data_validation(projects_validation)

        # Замораживаем первую строку
        ws_data.freeze_panes = 'A2'

        return workbook

    def generate_backup(self):
        """
        Генерация полного backup со всеми данными пользователей.

        Включает дополнительные поля для восстановления:
        - ID пользователя
        - Дата создания
        - Дата обновления
        - Статус (активен, одобрен, подтвержден)

        Returns:
            openpyxl.Workbook: Готовый workbook для сохранения
        """
        from apps.users.models import User

        workbook = openpyxl.Workbook()
        ws_data = workbook.active
        ws_data.title = "Backup"

        # Расширенные заголовки для backup
        backup_headers = [
            'ID', 'Email', 'ФИО', 'Роль', 'Должность', 'Телефон',
            'Проекты', 'Категория', 'Активен', 'Подтвержден', 'Одобрен',
            'Одобрен директором', 'Владелец компании', 'Полный доступ',
            'Дата создания', 'Дата обновления'
        ]

        # Заголовки
        for col_idx, header in enumerate(backup_headers, start=1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.CELL_BORDER

        # Устанавливаем ширину колонок
        ws_data.column_dimensions['A'].width = 8   # ID
        ws_data.column_dimensions['B'].width = 30  # Email
        ws_data.column_dimensions['C'].width = 25  # ФИО
        ws_data.column_dimensions['D'].width = 25  # Роль
        ws_data.column_dimensions['E'].width = 20  # Должность
        ws_data.column_dimensions['F'].width = 15  # Телефон
        ws_data.column_dimensions['G'].width = 40  # Проекты
        ws_data.column_dimensions['H'].width = 15  # Категория
        ws_data.column_dimensions['I'].width = 10  # Активен
        ws_data.column_dimensions['J'].width = 12  # Подтвержден
        ws_data.column_dimensions['K'].width = 10  # Одобрен
        ws_data.column_dimensions['L'].width = 18  # Одобрен директором
        ws_data.column_dimensions['M'].width = 16  # Владелец компании
        ws_data.column_dimensions['N'].width = 14  # Полный доступ
        ws_data.column_dimensions['O'].width = 18  # Дата создания
        ws_data.column_dimensions['P'].width = 18  # Дата обновления

        ws_data.row_dimensions[1].height = 30

        # Получаем пользователей
        users = User.objects.filter(
            company=self.company
        ).exclude(
            is_superuser=True
        ).select_related('company').prefetch_related('projects').order_by('created_at')

        # Заполняем данными
        for row_idx, user in enumerate(users, start=2):
            # ФИО
            full_name = f"{user.last_name} {user.first_name}"
            if user.middle_name:
                full_name += f" {user.middle_name}"

            # Проекты
            user_projects = user.projects.all().values_list('name', flat=True)
            projects_str = ', '.join(user_projects) if user_projects else ''

            # Данные строки
            row_data = [
                user.id,
                user.email,
                full_name.strip(),
                user.get_role_display(),
                user.position or '',
                user.phone or '',
                projects_str,
                user.get_role_category_display() if user.role_category else '',
                'Да' if user.is_active else 'Нет',
                'Да' if user.is_verified else 'Нет',
                'Да' if user.approved else 'Нет',
                'Да' if user.approved_by_director else 'Нет',
                'Да' if user.is_company_owner else 'Нет',
                'Да' if user.has_full_access else 'Нет',
                user.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                user.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name='Arial', size=10)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = self.CELL_BORDER

        # Замораживаем первую строку
        ws_data.freeze_panes = 'A2'

        # Добавляем информационный лист
        ws_info = workbook.create_sheet("Информация")
        ws_info.cell(row=1, column=1, value="ИНФОРМАЦИЯ О BACKUP").font = Font(size=14, bold=True)
        ws_info.cell(row=2, column=1, value=f"Компания: {self.company.name}")
        ws_info.cell(row=3, column=1, value=f"Дата создания backup: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws_info.cell(row=4, column=1, value=f"Количество пользователей: {users.count()}")
        ws_info.cell(row=5, column=1, value="")
        ws_info.cell(row=6, column=1, value="ВАЖНО:").font = Font(bold=True)
        ws_info.cell(row=7, column=1, value="Этот файл содержит полную информацию о пользователях.")
        ws_info.cell(row=8, column=1, value="Не используйте его для импорта (используйте Export для редактирования данных).")
        ws_info.cell(row=9, column=1, value="Файл предназначен для архивных целей и восстановления в случае необходимости.")

        ws_info.column_dimensions['A'].width = 80

        return workbook

    @staticmethod
    def get_role_mapping():
        """
        Создаёт словарь для конвертации русских названий ролей в коды.

        Поддерживает обратную совместимость: принимает как русские названия, так и коды.

        Returns:
            dict: Словарь вида:
                {
                    'Директор': 'DIRECTOR',
                    'DIRECTOR': 'DIRECTOR',  # для обратной совместимости
                    'Главный инженер': 'CHIEF_ENGINEER',
                    'CHIEF_ENGINEER': 'CHIEF_ENGINEER',
                    ...
                }

        Example:
            >>> mapping = PersonnelExcelHandler.get_role_mapping()
            >>> mapping['Директор']
            'DIRECTOR'
            >>> mapping['DIRECTOR']  # обратная совместимость
            'DIRECTOR'
        """
        from apps.users.models import User

        mapping = {}

        for role_code, role_label in User.Role.choices:
            if role_code != 'SUPERADMIN':
                # Маппинг русского названия -> код
                mapping[role_label] = role_code
                # Маппинг кода -> код (для обратной совместимости)
                mapping[role_code] = role_code

        return mapping

    @staticmethod
    def _is_valid_email(email):
        """
        Валидация email.

        Args:
            email (str): Email для проверки

        Returns:
            bool: True если email валиден
        """
        if not email:
            return False

        email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(email_regex, email) is not None


    def parse_import_file(self, file, mode='create'):
        """
        Парсинг и валидация Excel файла для импорта.

        Args:
            file: UploadedFile объект
            mode (str): Режим импорта - 'create' или 'update'

        Returns:
            dict: {
                'valid_rows': [{'email': ..., 'full_name': ..., ...}],
                'errors': [{'row': 5, 'errors': ['...']}]
            }
        """
        from apps.users.models import User

        try:
            workbook = openpyxl.load_workbook(file, data_only=True)
        except Exception as e:
            logger.error(f'Ошибка чтения Excel файла: {e}')
            return {
                'valid_rows': [],
                'errors': [{'row': 0, 'errors': [f'Не удалось прочитать файл: {str(e)}']}]
            }

        # Ищем лист "Данные"
        if "Данные" in workbook.sheetnames:
            sheet = workbook["Данные"]
        else:
            sheet = workbook.active

        valid_rows = []
        errors = []

        # Получаем список проектов компании для валидации
        company_projects_dict = {
            p.name.strip().lower(): p.id
            for p in self.company.projects.filter(is_active=True)
        }

        # Парсим строки (начиная со 2-й, т.к. 1-я — заголовок)
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Пропускаем пустые строки
            if not any(row):
                continue

            # Извлекаем значения колонок с безопасной обработкой типов
            # row[i] может быть str, int, float, None - приводим все к str
            email = str(row[0]).strip() if row[0] is not None else None
            full_name = str(row[1]).strip() if row[1] is not None else None
            role = str(row[2]).strip() if row[2] is not None else None
            position = str(row[3]).strip() if (len(row) > 3 and row[3] is not None) else ''
            phone = str(row[4]).strip() if (len(row) > 4 and row[4] is not None) else ''
            projects_str = str(row[5]).strip() if (len(row) > 5 and row[5] is not None) else ''

            row_errors = []

            # ===== ВАЛИДАЦИЯ ОБЯЗАТЕЛЬНЫХ ПОЛЕЙ =====

            # 1. Email
            if not email:
                row_errors.append('Email обязателен')
            elif not self._is_valid_email(email):
                row_errors.append('Неверный формат email')

            # 2. ФИО
            if not full_name:
                row_errors.append('ФИО обязательно')

            # 3. Роль
            if not role:
                row_errors.append('Роль обязательна')
            else:
                # Создаём маппинг русских названий к кодам (с поддержкой обратной совместимости)
                role_mapping = self.get_role_mapping()

                # Конвертируем русское название в код (если это русское название)
                # Если это уже код - он останется без изменений благодаря обратной совместимости
                if role in role_mapping:
                    role = role_mapping[role]  # 'Директор' -> 'DIRECTOR' или 'DIRECTOR' -> 'DIRECTOR'

                # Валидируем код роли
                if role not in [r[0] for r in User.Role.choices]:
                    row_errors.append(f'Неверная роль: {role}. Выберите роль из списка.')
                elif role == 'SUPERADMIN':
                    row_errors.append('Нельзя создавать пользователей с ролью SUPERADMIN через импорт')

            # ===== ВАЛИДАЦИЯ ДУБЛИКАТОВ EMAIL =====

            if email and self._is_valid_email(email):
                if mode == 'create':
                    # При создании проверяем, что email не существует
                    if User.objects.filter(email__iexact=email).exists():
                        row_errors.append(f'Email {email} уже существует в системе')

                elif mode == 'update':
                    # При обновлении проверяем, что пользователь существует в компании
                    if not User.objects.filter(email__iexact=email, company=self.company).exists():
                        row_errors.append(
                            f'Пользователь с email {email} не найден в вашей компании. '
                            f'Используйте режим "Массовое добавление" для создания новых пользователей.'
                        )

            # ===== ВАЛИДАЦИЯ НЕОБЯЗАТЕЛЬНЫХ ПОЛЕЙ =====

            # 4. Телефон (необязательное поле, любой формат разрешён)
            if phone:
                phone = str(phone).strip()
            else:
                phone = ''

            # 5. Проекты (если указаны)
            project_ids = []
            if projects_str:
                # Разбиваем по запятой и очищаем
                project_names = [p.strip() for p in projects_str.split(',') if p.strip()]

                for project_name in project_names:
                    project_name_lower = project_name.strip().lower()

                    if project_name_lower not in company_projects_dict:
                        row_errors.append(
                            f'Проект "{project_name}" не найден в вашей компании. '
                            f'Проверьте правильность написания.'
                        )
                    else:
                        project_ids.append(company_projects_dict[project_name_lower])

            # ===== ДОБАВЛЯЕМ РЕЗУЛЬТАТ =====

            if row_errors:
                errors.append({'row': idx, 'errors': row_errors})
            else:
                # Разбиваем ФИО на части
                name_parts = full_name.split()

                last_name = name_parts[0] if len(name_parts) > 0 else ''
                first_name = name_parts[1] if len(name_parts) > 1 else ''
                middle_name = name_parts[2] if len(name_parts) > 2 else ''

                valid_rows.append({
                    'email': email.lower(),  # Приводим к нижнему регистру
                    'last_name': last_name,
                    'first_name': first_name,
                    'middle_name': middle_name,
                    'role': role,
                    'position': position,
                    'phone': phone,
                    'project_ids': project_ids
                })

        logger.info(
            f'Парсинг завершен: {len(valid_rows)} валидных строк, '
            f'{len(errors)} ошибок (режим: {mode})'
        )

        return {
            'valid_rows': valid_rows,
            'errors': errors
        }
