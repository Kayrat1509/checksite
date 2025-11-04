"""
Обработчик Excel для импорта/экспорта подрядчиков.

Поддерживает три варианта работы:
1. Массовое добавление (создание новых подрядчиков)
2. Обновление данных (export → edit → import)
3. Backup (полная выгрузка с timestamp)

Особенности:
- Генерация постоянных паролей (не временных)
- Валидация email, телефонов, проектов
- Dropdown списки для проектов
- Отправка credentials на email
- Роль всегда 'CONTRACTOR'
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import re
import logging

logger = logging.getLogger(__name__)


class ContractorExcelHandler:
    """
    Класс для работы с Excel файлами подрядчиков.

    Методы:
    - generate_template_v2() — создать шаблон для импорта с dropdown
    - generate_export_v2() — экспорт текущих подрядчиков
    - generate_backup() — создать backup с timestamp
    - parse_import_file() — парсинг и валидация импортируемого файла
    """

    # Константы для Excel
    HEADER_FILL = PatternFill(start_color='FF6B35', end_color='FF6B35', fill_type='solid')  # Оранжевый для подрядчиков
    HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    CELL_BORDER = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Заголовки колонок для подрядчиков
    COLUMNS = {
        'email': 'Email*',
        'full_name': 'ФИО*',
        'contractor_company': 'Компания подрядчика*',
        'work_type': 'Вид работ',
        'phone': 'Телефон',
        'projects': 'Объекты'
    }

    # Ширина колонок (в символах)
    COLUMN_WIDTHS = {
        'A': 30,  # Email
        'B': 30,  # ФИО
        'C': 30,  # Компания подрядчика
        'D': 30,  # Вид работ
        'E': 18,  # Телефон
        'F': 40   # Объекты
    }

    def __init__(self, company):
        """
        Args:
            company: Company instance для изоляции данных
        """
        self.company = company

    def generate_template_v2(self):
        """
        Генерация шаблона Excel v2 для импорта подрядчиков.

        Структура:
        - Лист 1: "Инструкция" (ReadOnly) — описание процесса импорта
        - Лист 2: "Данные" (Editable) — таблица с данными и dropdown

        Returns:
            openpyxl.Workbook: Готовый workbook для сохранения
        """
        workbook = openpyxl.Workbook()

        # ===== ЛИСТ 1: ИНСТРУКЦИЯ =====
        ws_instruction = workbook.active
        ws_instruction.title = "Инструкция"

        instructions = [
            ("ИНСТРУКЦИЯ ПО ИМПОРТУ ПОДРЯДЧИКОВ", 16, True),
            ("", 11, False),
            ("1. ОБЯЗАТЕЛЬНЫЕ ПОЛЯ (отмечены звездочкой *):", 11, True),
            ("   - Email* — уникальный email подрядчика", 11, False),
            ("   - ФИО* — полное имя (Фамилия Имя Отчество или Фамилия Имя)", 11, False),
            ("   - Компания подрядчика* — название компании подрядчика", 11, False),
            ("", 11, False),
            ("2. НЕОБЯЗАТЕЛЬНЫЕ ПОЛЯ:", 11, True),
            ("   - Вид работ — описание видов выполняемых работ", 11, False),
            ("     Примеры: Электромонтажные работы, Сантехнические работы, Отделочные работы", 11, False),
            ("   - Телефон — контактный телефон", 11, False),
            ("     Форматы: +7XXXXXXXXXX или +7 (XXX) XXX-XX-XX", 11, False),
            ("   - Объекты — можно выбрать из выпадающего списка ИЛИ ввести несколько через запятую:", 11, False),
            ("     Пример 1: Выбрать из списка → Жилой комплекс Восход", 11, False),
            ("     Пример 2: Ввести несколько → Жилой комплекс Восход, ТЦ Запад, Школа №5", 11, False),
            ("     (при вводе нескольких объектов разделяйте их запятой)", 11, False),
            ("", 11, False),
            ("3. ВАЖНО:", 11, True),
            ("   - Email должен быть уникальным в системе", 11, False),
            ("   - При импорте будет сгенерирован постоянный пароль (12 символов)", 11, False),
            ("   - Данные для входа будут отправлены на указанный email", 11, False),
            ("   - Подрядчики будут привязаны к вашей компании автоматически", 11, False),
            ("   - Роль 'Подрядчик' устанавливается автоматически", 11, False),
            ("   - Рекомендуется сменить пароль в профиле после первого входа", 11, False),
            ("", 11, False),
            ("4. ПОРЯДОК РАБОТЫ:", 11, True),
            ("   1) Заполните таблицу на листе 'Данные'", 11, False),
            ("   2) Сохраните файл", 11, False),
            ("   3) Загрузите через кнопку 'Импорт' на странице Подрядчики", 11, False),
            ("   4) Дождитесь завершения импорта и проверьте статистику", 11, False),
            ("", 11, False),
            ("5. ПРИМЕР ЗАПОЛНЕНИЯ:", 11, True),
            ("   Email: ivanov@stroi.kz", 11, False),
            ("   ФИО: Иванов Иван Иванович", 11, False),
            ("   Компания подрядчика: ТОО \"СтройСервис\"", 11, False),
            ("   Вид работ: Электромонтажные работы", 11, False),
            ("   Телефон: +7 777 632 36 16", 11, False),
            ("   Объекты: Жилой комплекс \"Север\", ТЦ \"Запад\"", 11, False),
            ("", 11, False),
            (f"Дата создания шаблона: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 10, False),
            (f"Компания: {self.company.name}", 10, False),
        ]

        # Заполняем инструкцию
        for idx, (text, font_size, is_bold) in enumerate(instructions, start=1):
            cell = ws_instruction.cell(row=idx, column=1, value=text)
            cell.font = Font(name='Arial', size=font_size, bold=is_bold)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Устанавливаем ширину колонки A для инструкции
        ws_instruction.column_dimensions['A'].width = 100

        # ===== ЛИСТ 2: ДАННЫЕ =====
        ws_data = workbook.create_sheet("Данные")

        # Заголовки
        headers = list(self.COLUMNS.values())
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.CELL_BORDER

        # Устанавливаем ширину колонок
        for col_letter, width in self.COLUMN_WIDTHS.items():
            ws_data.column_dimensions[col_letter].width = width

        # Устанавливаем высоту заголовка
        ws_data.row_dimensions[1].height = 30

        # ===== DROPDOWN ДЛЯ ПРОЕКТОВ/ОБЪЕКТОВ =====
        # Получаем проекты компании (используем ТОТ ЖЕ метод что и у сотрудников)
        company_projects = self.company.projects.filter(is_active=True).values_list('name', flat=True)

        if company_projects:
            # Ограничение Excel: формула не может быть длиннее 255 символов
            project_list = ','.join(company_projects)

            if len(project_list) <= 250:
                # Если список проектов короткий, используем прямую формулу (как у сотрудников)
                projects_formula = f'"{project_list}"'

                projects_validation = DataValidation(
                    type="list",
                    formula1=projects_formula,
                    allow_blank=True,
                    showErrorMessage=False,  # Разрешаем свободный ввод для множественного выбора
                    showInputMessage=True,
                    promptTitle='Выбор объектов',
                    prompt='Выберите один объект из списка или введите несколько через запятую (например: Объект 1, Объект 2)'
                )
                projects_validation.add('F2:F1000')  # Применяем к колонке "Объекты" (F)
                ws_data.add_data_validation(projects_validation)
            else:
                # Если список длинный, создаем скрытый лист с проектами
                ws_projects = workbook.create_sheet("__Объекты__")

                for idx, project_name in enumerate(company_projects, start=1):
                    ws_projects.cell(row=idx, column=1, value=project_name)

                # Скрываем лист
                ws_projects.sheet_state = 'hidden'

                # Создаем валидацию с reference на скрытый лист
                projects_validation = DataValidation(
                    type="list",
                    formula1=f"'__Объекты__'!$A$1:$A${len(company_projects)}",
                    allow_blank=True,
                    showErrorMessage=False,  # Разрешаем свободный ввод для множественного выбора
                    showInputMessage=True,
                    promptTitle='Выбор объектов',
                    prompt='Выберите один объект из списка или введите несколько через запятую'
                )
                projects_validation.add('F2:F1000')
                ws_data.add_data_validation(projects_validation)

        # Добавляем примеры данных (2 строки)
        examples = [
            {
                'email': 'ivanov@stroi.kz',
                'full_name': 'Иванов Иван Иванович',
                'contractor_company': 'ТОО "СтройСервис"',
                'work_type': 'Электромонтажные работы',
                'phone': '+7 777 632 36 16',
                'projects': '' if not company_projects else list(company_projects)[0] if len(company_projects) == 1 else f"{list(company_projects)[0]}, {list(company_projects)[1] if len(company_projects) > 1 else ''}"
            },
            {
                'email': 'petrov@build.kz',
                'full_name': 'Петров Петр',
                'contractor_company': 'ИП "ПетровСтрой"',
                'work_type': 'Сантехнические работы',
                'phone': '+7 777 987 65 43',
                'projects': '' if not company_projects else list(company_projects)[0]
            }
        ]

        for row_idx, example in enumerate(examples, start=2):
            for col_idx, field in enumerate(self.COLUMNS.keys(), start=1):
                cell = ws_data.cell(row=row_idx, column=col_idx, value=example.get(field, ''))
                cell.border = self.CELL_BORDER
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        return workbook

    def generate_export_v2(self):
        """
        Экспорт текущих подрядчиков компании в Excel v2.

        Создает файл с текущими данными для редактирования и последующего
        реимпорта в режиме update.

        Returns:
            openpyxl.Workbook: Готовый workbook для сохранения
        """
        from apps.users.models import User

        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = "Подрядчики"

        # Заголовки
        headers = list(self.COLUMNS.values())
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.CELL_BORDER

        # Устанавливаем ширину колонок
        for col_letter, width in self.COLUMN_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width

        # Устанавливаем высоту заголовка
        ws.row_dimensions[1].height = 30

        # ===== DROPDOWN ДЛЯ ПРОЕКТОВ =====
        # Используем ТОТ ЖЕ метод что и у сотрудников
        company_projects = self.company.projects.filter(is_active=True).values_list('name', flat=True)

        if company_projects:
            # Ограничение Excel: формула не может быть длиннее 255 символов
            project_list = ','.join(company_projects)

            if len(project_list) <= 250:
                # Если список проектов короткий, используем прямую формулу
                projects_formula = f'"{project_list}"'

                projects_validation = DataValidation(
                    type="list",
                    formula1=projects_formula,
                    allow_blank=True,
                    showErrorMessage=False,
                    showInputMessage=True,
                    promptTitle='Выбор объектов',
                    prompt='Выберите один объект из списка или введите несколько через запятую'
                )
                projects_validation.add('F2:F1000')
                ws.add_data_validation(projects_validation)
            else:
                # Если список длинный, создаем скрытый лист
                ws_projects = workbook.create_sheet("__Объекты__")

                for idx, project_name in enumerate(company_projects, start=1):
                    ws_projects.cell(row=idx, column=1, value=project_name)

                ws_projects.sheet_state = 'hidden'

                projects_validation = DataValidation(
                    type="list",
                    formula1=f"'__Объекты__'!$A$1:$A${len(company_projects)}",
                    allow_blank=True,
                    showErrorMessage=False,
                    showInputMessage=True,
                    promptTitle='Выбор объектов',
                    prompt='Выберите один объект из списка или введите несколько через запятую'
                )
                projects_validation.add('F2:F1000')
                ws.add_data_validation(projects_validation)

        # ===== ДАННЫЕ ПОДРЯДЧИКОВ =====
        # Получаем только подрядчиков компании (не удаленных)
        contractors = User.objects.filter(
            company=self.company,
            role='CONTRACTOR',
            is_deleted=False
        ).prefetch_related('projects')

        for row_idx, contractor in enumerate(contractors, start=2):
            # Формируем список проектов через запятую
            project_names = ', '.join([p.name for p in contractor.projects.all()])

            # Формируем ФИО (Фамилия Имя Отчество)
            full_name_parts = [contractor.last_name, contractor.first_name]
            if contractor.middle_name:
                full_name_parts.append(contractor.middle_name)
            full_name = ' '.join(full_name_parts)

            row_data = {
                'email': contractor.email,
                'full_name': full_name,
                'contractor_company': contractor.external_company_name or '',
                'work_type': contractor.work_type or '',
                'phone': contractor.phone or '',
                'projects': project_names
            }

            for col_idx, field in enumerate(self.COLUMNS.keys(), start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(field, ''))
                cell.border = self.CELL_BORDER
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        return workbook

    def generate_backup(self):
        """
        Создание полного backup подрядчиков с timestamp.

        Включает все данные, включая статусы и даты для архивных целей.

        Returns:
            openpyxl.Workbook: Готовый workbook для сохранения
        """
        from apps.users.models import User

        workbook = openpyxl.Workbook()

        # ===== ЛИСТ 1: ИНФОРМАЦИЯ О BACKUP =====
        ws_info = workbook.active
        ws_info.title = "Информация"

        backup_info = [
            ("BACKUP ПОДРЯДЧИКОВ", 16, True),
            ("", 11, False),
            (f"Дата создания: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 11, False),
            (f"Компания: {self.company.name}", 11, False),
            ("", 11, False),
            ("Содержание:", 11, True),
            ("- Лист 'Подрядчики': все подрядчики компании (активные и архивированные)", 11, False),
            ("- Включает расширенные данные: ID, статусы, даты", 11, False),
            ("", 11, False),
            ("Назначение:", 11, True),
            ("- Архивное хранение данных", 11, False),
            ("- Восстановление в случае потери данных", 11, False),
            ("- Анализ и отчетность", 11, False),
        ]

        for idx, (text, font_size, is_bold) in enumerate(backup_info, start=1):
            cell = ws_info.cell(row=idx, column=1, value=text)
            cell.font = Font(name='Arial', size=font_size, bold=is_bold)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

        ws_info.column_dimensions['A'].width = 80

        # ===== ЛИСТ 2: ДАННЫЕ ПОДРЯДЧИКОВ =====
        ws_data = workbook.create_sheet("Подрядчики")

        # Расширенные заголовки для backup
        backup_columns = {
            'id': 'ID',
            'email': 'Email',
            'full_name': 'ФИО',
            'contractor_company': 'Компания подрядчика',
            'work_type': 'Вид работ',
            'phone': 'Телефон',
            'projects': 'Объекты',
            'is_active': 'Активен',
            'is_deleted': 'В корзине',
            'created_at': 'Дата создания',
            'updated_at': 'Дата обновления'
        }

        # Заголовки
        for col_idx, header in enumerate(backup_columns.values(), start=1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.CELL_BORDER

        # Устанавливаем ширину колонок
        column_widths = [6, 30, 30, 30, 30, 18, 40, 10, 12, 20, 20]
        for idx, width in enumerate(column_widths, start=1):
            ws_data.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

        # Устанавливаем высоту заголовка
        ws_data.row_dimensions[1].height = 30

        # ===== ДАННЫЕ ПОДРЯДЧИКОВ (ВСЕ, включая архивированных) =====
        contractors = User.objects.filter(
            company=self.company,
            role='CONTRACTOR'
        ).prefetch_related('projects')

        for row_idx, contractor in enumerate(contractors, start=2):
            # Формируем список проектов через запятую
            project_names = ', '.join([p.name for p in contractor.projects.all()])

            # Формируем ФИО (Фамилия Имя Отчество)
            full_name_parts = [contractor.last_name, contractor.first_name]
            if contractor.middle_name:
                full_name_parts.append(contractor.middle_name)
            full_name = ' '.join(full_name_parts)

            row_data = {
                'id': contractor.id,
                'email': contractor.email,
                'full_name': full_name,
                'contractor_company': contractor.external_company_name or '',
                'work_type': contractor.work_type or '',
                'phone': contractor.phone or '',
                'projects': project_names,
                'is_active': 'Да' if contractor.is_active else 'Нет',
                'is_deleted': 'Да' if contractor.is_deleted else 'Нет',
                'created_at': contractor.created_at.strftime('%Y-%m-%d %H:%M:%S') if contractor.created_at else '',
                'updated_at': contractor.updated_at.strftime('%Y-%m-%d %H:%M:%S') if contractor.updated_at else ''
            }

            for col_idx, field in enumerate(backup_columns.keys(), start=1):
                cell = ws_data.cell(row=row_idx, column=col_idx, value=row_data.get(field, ''))
                cell.border = self.CELL_BORDER
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        return workbook

    def parse_import_file(self, file, mode='create'):
        """
        Парсинг и валидация Excel файла для импорта подрядчиков.

        Args:
            file: Upload file object (Excel файл)
            mode: 'create' или 'update'

        Returns:
            dict: {
                'valid_rows': [список валидных строк],
                'errors': [список ошибок валидации]
            }
        """
        from apps.users.models import User
        from apps.projects.models import Project

        try:
            workbook = openpyxl.load_workbook(file)
            ws = workbook['Данные'] if 'Данные' in workbook.sheetnames else workbook.active
        except Exception as e:
            logger.error(f"Ошибка при чтении Excel файла: {str(e)}")
            return {
                'valid_rows': [],
                'errors': [{'row': 0, 'errors': [f'Не удалось прочитать файл: {str(e)}']}]
            }

        valid_rows = []
        errors = []

        # Начинаем с 2-й строки (1-я строка — заголовки)
        for row_idx in range(2, ws.max_row + 1):
            row_errors = []

            # Читаем данные из ячеек (новая структура: Email, ФИО, Компания, Вид работ, Телефон, Объекты)
            email = ws.cell(row=row_idx, column=1).value
            full_name = ws.cell(row=row_idx, column=2).value
            contractor_company = ws.cell(row=row_idx, column=3).value
            work_type = ws.cell(row=row_idx, column=4).value
            phone = ws.cell(row=row_idx, column=5).value
            projects_str = ws.cell(row=row_idx, column=6).value

            # Пропускаем пустые строки
            if not any([email, full_name, contractor_company]):
                continue

            # === ВАЛИДАЦИЯ ОБЯЗАТЕЛЬНЫХ ПОЛЕЙ ===

            # Email
            if not email or not isinstance(email, str):
                row_errors.append('Email обязателен')
            else:
                email = email.strip().lower()
                # Проверка формата email
                if not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
                    row_errors.append('Некорректный формат email')
                else:
                    # Проверка уникальности email
                    if mode == 'create':
                        if User.objects.filter(email__iexact=email).exists():
                            row_errors.append(f'Email {email} уже существует в системе')
                    elif mode == 'update':
                        if not User.objects.filter(email__iexact=email, company=self.company, role='CONTRACTOR').exists():
                            row_errors.append(f'Подрядчик с email {email} не найден')

            # ФИО (парсим на фамилию, имя, отчество)
            if not full_name or not isinstance(full_name, str) or not full_name.strip():
                row_errors.append('ФИО обязательно')
                last_name = ''
                first_name = ''
                middle_name = ''
            else:
                full_name = full_name.strip()
                # Разбиваем ФИО на части: "Иванов Иван Иванович" → ['Иванов', 'Иван', 'Иванович']
                name_parts = full_name.split()

                if len(name_parts) < 2:
                    row_errors.append('ФИО должно содержать минимум Фамилию и Имя (например: Иванов Иван)')
                    last_name = name_parts[0] if len(name_parts) > 0 else ''
                    first_name = ''
                    middle_name = ''
                else:
                    last_name = name_parts[0]
                    first_name = name_parts[1]
                    middle_name = ' '.join(name_parts[2:]) if len(name_parts) > 2 else ''

            # Компания подрядчика
            if not contractor_company or not isinstance(contractor_company, str) or not contractor_company.strip():
                row_errors.append('Компания подрядчика обязательна')
            else:
                contractor_company = contractor_company.strip()

            # === ВАЛИДАЦИЯ НЕОБЯЗАТЕЛЬНЫХ ПОЛЕЙ ===

            # Вид работ (необязательное поле)
            if work_type:
                work_type = str(work_type).strip()
            else:
                work_type = ''

            # Телефон (необязательное поле, любой формат разрешён)
            if phone:
                phone = str(phone).strip()
            else:
                phone = ''

            # Проекты/Объекты
            project_ids = []
            if projects_str:
                projects_str = str(projects_str).strip()
                # Разбиваем по запятой
                project_names = [p.strip() for p in projects_str.split(',') if p.strip()]

                for project_name in project_names:
                    try:
                        project = Project.objects.get(
                            name__iexact=project_name,
                            company=self.company,
                            is_active=True
                        )
                        project_ids.append(project.id)
                    except Project.DoesNotExist:
                        row_errors.append(f'Проект "{project_name}" не найден')
                    except Project.MultipleObjectsReturned:
                        row_errors.append(f'Найдено несколько проектов с названием "{project_name}"')

            # Если есть ошибки валидации, добавляем в список ошибок
            if row_errors:
                errors.append({
                    'row': row_idx,
                    'errors': row_errors
                })
            else:
                # Добавляем валидную строку
                valid_rows.append({
                    'email': email,
                    'last_name': last_name,
                    'first_name': first_name,
                    'middle_name': middle_name,
                    'contractor_company': contractor_company,
                    'work_type': work_type,
                    'phone': phone,
                    'project_ids': project_ids
                })

        return {
            'valid_rows': valid_rows,
            'errors': errors
        }
